"""Impor kotak masuk (Excel milik pemilik) → buku besar. APPEND-ONLY.

Aturan emas: kotak masuk hanya boleh MENAMBAH. Ia tidak pernah bisa mengurangi.

    kunci baru               -> TAMBAH
    kunci ada, nilai sama    -> lewati
    kunci ada, nilai beda    -> TOLAK + catat konflik (buku besar menang)
    kunci ada di buku besar,
      hilang dari workbook   -> lewati (buku besar tetap menyimpannya)
    kunci di periode terkunci-> TOLAK, walau nilainya baru
    kolom berangka tak
      terpetakan             -> GAGAL KERAS, hentikan impor

Kolom `Total ...` di workbook TIDAK PERNAH dibaca. Ia terbukti tidak dapat
dipercaya: Overview Maret menulis 43.160 kg padahal totalnya 50.189 kg, karena
SOD RS dan Pick Up tidak ikut dijumlahkan. Nilai selalu dihitung ulang dari
kolom kendaraan.

Bawaan: DRY-RUN. Tidak ada yang ditulis sampai Anda menambahkan --terapkan.

    python import_inbox.py                      # pratinjau kedua dataset
    python import_inbox.py --dataset timbulan   # satu dataset saja
    python import_inbox.py --terapkan           # tulis ke buku besar
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _ledger import (
    ABAIKAN,
    fmt,
    is_frozen,
    ledger_dir,
    load_freeze,
    load_peta,
    match_kategori,
    read_rows,
    same,
    today_iso,
    write_conflict_report,
    write_md_mirror,
    write_rows,
)
from _utils import find_source

try:
    import openpyxl
except ImportError:  # pragma: no cover
    print("[import_inbox] butuh openpyxl: python -m pip install openpyxl", file=sys.stderr)
    raise

TIMBULAN_XLSX = "Copy of Total Timbulan Sampah 2026  (Bulanan).xlsx"
TRAFFIC_XLSX = "Kecelakaan Lalu Lintas.xlsx"

CATEGORY_COLS = ["organik_anorganik", "sisa_makanan", "lingkungan", "aset"]
TIMBULAN_FIELDS = ["tanggal", "hari", *[f"{c}_kg" for c in CATEGORY_COLS], "ipdn_kg",
                   "catatan", "quality_flag", "sumber", "dicatat_pada"]
TRAFFIC_FIELDS = ["tahun", "bulan", "jenis", "jumlah", "sumber", "dicatat_pada"]
TRAFFIC_DETAIL_FIELDS = ["tahun", "no", "bulan", "jenis", "lokasi_id", "lokasi_label",
                         "jumlah", "catatan", "sumber", "dicatat_pada"]

DAY_ID = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
YEAR_SHEET = re.compile(r"^\d{4}$")


class KolomTakDikenal(SystemExit):
    pass


# ── Pembacaan kotak masuk: timbulan ──────────────────────────────────────


def _labels_row1(ws) -> list[str]:
    """Baris 1 berisi label grup di sel merge. Isi ke kanan sampai label baru."""
    out, cur = [], ""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v not in (None, ""):
            cur = str(v).strip()
        out.append(cur)
    return out


def baca_timbulan(path: Path, peta: list[dict]) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    wb_raw = openpyxl.load_workbook(path, data_only=False)
    rows: list[dict] = []
    tak_dikenal: dict[str, float] = {}
    rumus_kosong: list[str] = []

    for ws in wb.worksheets:
        ws_raw = wb_raw[ws.title]
        labels = _labels_row1(ws)
        if "tanggal" not in [l.lower() for l in labels]:
            continue  # Overview, Rekapan, Dashboard — bukan sheet harian
        col_tgl = [i for i, l in enumerate(labels) if l.lower() == "tanggal"][0] + 1

        for r in range(2, ws.max_row + 1):
            tgl = ws.cell(r, col_tgl).value
            if isinstance(tgl, dt.datetime):
                tgl = tgl.date()
            if not isinstance(tgl, dt.date):
                continue
            iso = tgl.isoformat()

            cats = {c: 0.0 for c in CATEGORY_COLS}
            ipdn = 0.0
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                label = labels[c - 1]

                # Rumus tanpa nilai tersimpan akan terbaca kosong -> diam-diam nol.
                if v is None:
                    raw = ws_raw.cell(r, c).value
                    if isinstance(raw, str) and raw.startswith("="):
                        kat = match_kategori(label, iso, peta)
                        if kat is not None and kat != ABAIKAN and len(rumus_kosong) < 6:
                            rumus_kosong.append(f"{ws.title}!{ws.cell(r, c).coordinate}  ({raw})")
                    continue

                if not isinstance(v, (int, float)) or v == 0:
                    continue
                kat = match_kategori(label, iso, peta)
                if kat is None:
                    kunci = f"{ws.title}!{label or '(tanpa label)'}"
                    tak_dikenal[kunci] = tak_dikenal.get(kunci, 0) + float(v)
                elif kat == ABAIKAN:
                    continue
                elif kat == "ipdn":
                    ipdn += float(v)
                else:
                    cats[kat] += float(v)

            if sum(cats.values()) + ipdn == 0:
                continue
            rows.append({
                "tanggal": iso,
                "hari": DAY_ID[tgl.weekday()],
                **{f"{c}_kg": fmt(cats[c]) for c in CATEGORY_COLS},
                "ipdn_kg": fmt(ipdn),
                "catatan": "", "quality_flag": "",
                "sumber": path.name, "dicatat_pada": today_iso(),
            })
    wb.close()
    wb_raw.close()

    if rumus_kosong:
        detail = "\n".join(f"    {x}" for x in rumus_kosong)
        raise KolomTakDikenal(
            f"\n[import_inbox] GAGAL KERAS — sel berisi RUMUS tanpa nilai tersimpan:\n{detail}\n"
            f"  Sel semacam ini terbaca KOSONG, sehingga beratnya akan hilang diam-diam.\n"
            f"  Buka '{path.name}' di Excel, simpan (Ctrl+S), lalu ulangi impor.\n"
            f"  (Terjadi bila workbook pernah ditulis ulang oleh skrip, bukan oleh Excel.)\n"
        )

    if tak_dikenal:
        detail = "\n".join(f"    {k}  ({v:,.0f} kg)" for k, v in sorted(tak_dikenal.items()))
        raise KolomTakDikenal(
            f"\n[import_inbox] GAGAL KERAS — kolom berangka yang belum dipetakan:\n{detail}\n"
            f"  Angka di atas TIDAK akan dibuang diam-diam. Impor dihentikan.\n"
            f"  Tambahkan polanya ke data/_ledger/_peta_kolom.csv, lalu ulangi.\n"
        )
    return rows


# ── Pembacaan kotak masuk: kecelakaan ────────────────────────────────────


def baca_traffic(path: Path) -> tuple[list[dict], list[dict]]:
    from parse_traffic_accidents_xlsx import (
        LOCATION_ID_MAP, MONTHS_ID, TYPE_MAP, _match_type_id, parse_year_sheet,
    )

    wb = openpyxl.load_workbook(path, data_only=True)
    monthly: list[dict] = []
    detail: list[dict] = []

    for name in wb.sheetnames:
        if not YEAR_SHEET.match(name):
            continue
        year = int(name)
        ws = wb[name]
        for m in parse_year_sheet(ws, year)["monthly"]:
            for jenis, jml in m["by_type"].items():
                monthly.append({"tahun": year, "bulan": m["month"], "jenis": jenis,
                                "jumlah": jml, "sumber": path.name, "dicatat_pada": today_iso()})

        cur_month, no = None, 0
        for r in range(4, ws.max_row + 1):
            c16, c17, c18, c19 = (ws.cell(r, i).value for i in (16, 17, 18, 19))
            if isinstance(c16, str) and c16.strip() in MONTHS_ID and c17 is None:
                cur_month = f"{year}-{MONTHS_ID.index(c16.strip()) + 1:02d}"
                continue
            if isinstance(c16, (int, float)) and c17 and c18:
                no += 1
                jenis_raw = str(c17).strip()
                lokasi_raw = str(c18).strip()
                tid = _match_type_id(jenis_raw)
                catatan = jenis_raw if (tid is None or jenis_raw.lower() not in
                                        [t.lower() for t in TYPE_MAP]) else ""
                detail.append({
                    "tahun": year, "no": no, "bulan": cur_month or f"{year}-01",
                    "jenis": tid or "", "lokasi_id": LOCATION_ID_MAP.get(
                        lokasi_raw, lokasi_raw.lower().replace(" ", "-")),
                    "lokasi_label": lokasi_raw,
                    "jumlah": int(c19) if isinstance(c19, (int, float)) else 1,
                    "catatan": catatan, "sumber": path.name, "dicatat_pada": today_iso(),
                })
    wb.close()
    return monthly, detail


# ── Merge append-only ────────────────────────────────────────────────────


def merge(dataset: str, led: Path, nama_csv: str, fields: list[str],
          kunci: list[str], nilai: list[str], masuk: list[dict],
          bulan_dari: callable) -> dict:
    lama = read_rows(led / nama_csv)
    idx = {tuple(r[k] for k in kunci): r for r in lama}
    freeze = load_freeze(led)

    tambah, konflik, sama_, terkunci = [], [], [], []
    for r in masuk:
        key = tuple(str(r[k]) for k in kunci)
        bulan = bulan_dari(r)
        f = is_frozen(freeze, dataset, bulan)
        if f:
            terkunci.append((key, bulan))
            continue
        if key not in idx:
            tambah.append(r)
            continue
        old = idx[key]
        beda = _beda(old, r, nilai)
        if beda:
            konflik.append((key, [(n, old.get(n), r.get(n)) for n in beda]))
        else:
            sama_.append(key)

    kunci_masuk = {tuple(str(r[x]) for x in kunci) for r in masuk}
    hilang = [k for k in idx if k not in kunci_masuk]
    return {"lama": lama, "tambah": tambah, "konflik": konflik, "sama": sama_,
            "terkunci": terkunci, "hilang": hilang, "fields": fields, "kunci": kunci,
            "nama_csv": nama_csv}


def _numeric(a, b) -> bool:
    try:
        float(a if a not in (None, "") else 0)
        float(b if b not in (None, "") else 0)
        return True
    except (TypeError, ValueError):
        return False


def _beda(old: dict, baru: dict, nilai: list[str]) -> list[str]:
    """Kolom mana yang berbeda. Angka dibandingkan sebagai angka, sisanya sebagai teks."""
    out = []
    for n in nilai:
        o, b = old.get(n), baru.get(n)
        if _numeric(o, b):
            if not same(o, b):
                out.append(n)
        elif str(o or "").strip() != str(b or "").strip():
            out.append(n)
    return out


def terapkan(led: Path, hasil: dict, urut: callable) -> None:
    rows = hasil["lama"] + hasil["tambah"]
    rows.sort(key=urut)
    write_rows(led / hasil["nama_csv"], hasil["fields"], rows)


# ── Laporan ──────────────────────────────────────────────────────────────


def lapor(judul: str, h: dict) -> list[str]:
    baris = []
    print(f"\n── {judul} ──")
    print(f"  AKAN DITAMBAH  {len(h['tambah'])} baris")
    for r in h["tambah"][:10]:
        print(f"      + {' | '.join(str(r[k]) for k in h['kunci'])}")
    if len(h["tambah"]) > 10:
        print(f"      ... +{len(h['tambah']) - 10} lagi")

    print(f"  KONFLIK        {len(h['konflik'])} baris (DITOLAK — buku besar menang)")
    for key, beda in h["konflik"][:10]:
        for n, o, b in beda:
            msg = f"{' | '.join(key)} · {n}: buku besar {o} vs workbook {b}"
            print(f"      ! {msg}")
            baris.append(msg)

    print(f"  TERKUNCI       {len(h['terkunci'])} baris (periode dikunci, diabaikan)")
    print(f"  SAMA           {len(h['sama'])} baris")
    print(f"  HANYA DI BUKU BESAR  {len(h['hilang'])} baris (tetap disimpan, tidak dihapus)")
    return baris


def main() -> int:
    ap = argparse.ArgumentParser(description="Impor kotak masuk ke buku besar (append-only)")
    ap.add_argument("--dataset", choices=["timbulan", "kecelakaan", "semua"], default="semua")
    ap.add_argument("--ledger", default=None)
    ap.add_argument("--timbulan-source", default=None)
    ap.add_argument("--traffic-source", default=None)
    ap.add_argument("--terapkan", action="store_true", help="Tulis ke buku besar (bawaan: dry-run)")
    args = ap.parse_args()

    led = ledger_dir(args.ledger)
    if not (led / "_peta_kolom.csv").exists():
        raise SystemExit(f"[import_inbox] buku besar belum disemai di {led}. Jalankan seed_ledger.py.")
    peta = load_peta(led)

    mode = "TERAPKAN" if args.terapkan else "DRY-RUN (tidak ada yang ditulis)"
    print(f"[import_inbox] mode: {mode}")
    print(f"[import_inbox] buku besar: {led}")

    konflik_semua: dict[str, list[str]] = {}

    if args.dataset in ("timbulan", "semua"):
        src = Path(args.timbulan_source) if args.timbulan_source else find_source(TIMBULAN_XLSX)
        if not src or not src.exists():
            raise SystemExit(f"[import_inbox] sumber timbulan '{TIMBULAN_XLSX}' tidak ditemukan.")
        print(f"\n[timbulan] kotak masuk: {src}")
        masuk = baca_timbulan(src, peta)
        h = merge("timbulan", led, "timbulan.csv", TIMBULAN_FIELDS, ["tanggal"],
                  [f"{c}_kg" for c in CATEGORY_COLS] + ["ipdn_kg"], masuk,
                  lambda r: r["tanggal"][:7])
        konflik_semua["Timbulan - konflik ditolak"] = lapor("TIMBULAN", h)
        if args.terapkan and h["tambah"]:
            terapkan(led, h, lambda r: r["tanggal"])
            write_md_mirror(led / "timbulan.csv", led / "timbulan.md",
                            "Buku Besar — Timbulan Sampah",
                            "Nilai turunan (unpad_kg, total_kg) tidak disimpan; parser menghitungnya.")
            print(f"  -> {len(h['tambah'])} baris ditulis ke buku besar")

    if args.dataset in ("kecelakaan", "semua"):
        src = Path(args.traffic_source) if args.traffic_source else find_source(TRAFFIC_XLSX)
        if not src or not src.exists():
            raise SystemExit(f"[import_inbox] sumber kecelakaan '{TRAFFIC_XLSX}' tidak ditemukan.")
        print(f"\n[kecelakaan] kotak masuk: {src}")
        m_rows, d_rows = baca_traffic(src)

        hm = merge("traffic_accidents", led, "traffic_accidents.csv", TRAFFIC_FIELDS,
                   ["tahun", "bulan", "jenis"], ["jumlah"], m_rows, lambda r: r["bulan"])
        konflik_semua["Kecelakaan (bulanan) - konflik ditolak"] = lapor("KECELAKAAN - bulanan", hm)

        hd = merge("traffic_accidents", led, "traffic_accidents_detail.csv", TRAFFIC_DETAIL_FIELDS,
                   ["tahun", "no"], ["bulan", "jenis", "lokasi_id", "jumlah"], d_rows,
                   lambda r: r["bulan"])
        konflik_semua["Kecelakaan (detail lokasi) - konflik ditolak"] = lapor("KECELAKAAN - detail lokasi", hd)

        if args.terapkan and (hm["tambah"] or hd["tambah"]):
            if hm["tambah"]:
                terapkan(led, hm, lambda r: (str(r["tahun"]), str(r["bulan"]), str(r["jenis"])))
                write_md_mirror(led / "traffic_accidents.csv", led / "traffic_accidents.md",
                                "Buku Besar — Kecelakaan Lalu Lintas",
                                "Satu baris = satu (tahun, bulan, jenis). Total dihitung parser.")
            if hd["tambah"]:
                terapkan(led, hd, lambda r: (str(r["tahun"]), int(r["no"])))
            print(f"  -> {len(hm['tambah'])} baris bulanan + {len(hd['tambah'])} baris detail ditulis")

    if args.terapkan:
        path = write_conflict_report(led, today_iso(), konflik_semua)
        if path:
            print(f"\n[import_inbox] laporan konflik: {path}")
    else:
        print("\n[import_inbox] DRY-RUN selesai. TIDAK ADA yang ditulis — termasuk laporan konflik.")
        print("[import_inbox] Kalau isinya benar, ulangi dengan --terapkan")
    return 0


if __name__ == "__main__":
    sys.exit(main())
