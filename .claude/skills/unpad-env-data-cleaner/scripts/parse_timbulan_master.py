"""Bangun timbulan.json dari workbook induk 'Timbulan Sampah 2026 (MASTER).xlsx'.

Menggantikan parse_timbulan.py (yang membaca MD lama, skema 3-kategori, dan sudah
tertinggal dari dashboard). Workbook induk adalah SUMBER KEBENARAN tunggal:
setiap bulan baru ditambahkan sebagai baris di sheet 'Harian', tanpa pernah
menghapus baris lama. Karena itu rebuild penuh selalu aman.

Sheet yang dibaca:
  Harian            tanggal, hari, 4 kolom kategori kg, ipdn_kg, catatan, quality_flag
  Bulan             month (YYYY-MM), label, days_in_month
  Kategori          id, label, color_key, source
  Kendaraan         id, label, operator, tare_kg, note
  Tare Kontainer    kontainer, tare_kg
  Catatan Kualitas  severity, message

Nilai turunan TIDAK disimpan di workbook, dihitung di sini:
  unpad_kg = organik_anorganik + sisa_makanan + lingkungan + aset
  total_kg = unpad_kg + ipdn_kg

Efek samping yang disengaja: menulis snapshot teks sumber ke
<out>/_sources/timbulan_master.csv agar git menyimpan riwayat sumbernya
(workbook .xlsx sendiri di-gitignore, jadi tidak pernah masuk repo).

Butuh openpyxl (satu-satunya parser yang tidak stdlib-only, sama seperti
update_timbulan_from_xlsx.py sebelumnya).
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _utils import base_dataset, find_project_root, resolve_output, write_json

try:
    import openpyxl
except ImportError:  # pragma: no cover
    print("[parse_timbulan_master] butuh openpyxl: python -m pip install openpyxl", file=sys.stderr)
    raise

DATASET_ID = "timbulan"
MASTER_NAME = "Timbulan Sampah 2026 (MASTER).xlsx"
DAY_ID = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
CATEGORY_COLS = ["organik_anorganik", "sisa_makanan", "lingkungan", "aset"]


def locate_master(explicit: str | None) -> Path:
    """Cari workbook induk. Gagal keras kalau tidak ketemu — jangan pernah
    menghasilkan output kosong yang menimpa data bagus."""
    if explicit:
        p = Path(explicit)
        if not p.exists():
            raise SystemExit(f"[parse_timbulan_master] --source tidak ditemukan: {p}")
        return p

    env = os.environ.get("UNPAD_TIMBULAN_MASTER")
    if env and Path(env).exists():
        return Path(env)

    project = find_project_root()
    candidates = [
        project / MASTER_NAME,
        project.parent / "Data dan Pengetahuan" / MASTER_NAME,
        project.parent.parent / "Data dan Pengetahuan" / MASTER_NAME,
    ]
    for c in candidates:
        if c.exists():
            return c

    tried = "\n  ".join(str(c) for c in candidates)
    raise SystemExit(
        f"[parse_timbulan_master] workbook induk '{MASTER_NAME}' tidak ditemukan.\n"
        f"Dicari di:\n  {tried}\n"
        f"Set env UNPAD_TIMBULAN_MASTER atau pakai --source <path>.\n"
        f"ABORT — timbulan.json TIDAK diubah."
    )


def num(v) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def rows(ws) -> list[dict]:
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(it)]
    out = []
    for r in it:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append(dict(zip(headers, r)))
    return out


def read_daily(ws) -> list[dict]:
    entries: list[dict] = []
    seen: set[str] = set()
    for r in rows(ws):
        tgl = r.get("tanggal")
        if isinstance(tgl, dt.datetime):
            tgl = tgl.date()
        if not isinstance(tgl, dt.date):
            continue
        iso = tgl.isoformat()
        if iso in seen:
            raise SystemExit(f"[parse_timbulan_master] tanggal ganda di sheet Harian: {iso}. ABORT.")
        seen.add(iso)

        cats = {c: round(num(r.get(f"{c}_kg")), 3) for c in CATEGORY_COLS}
        ipdn = round(num(r.get("ipdn_kg")), 3)
        unpad = round(sum(cats.values()), 3)
        total = round(unpad + ipdn, 3)
        if total == 0:
            continue

        note = r.get("catatan")
        qf = r.get("quality_flag")
        entries.append({
            "date": iso,
            "day_of_week": DAY_ID[tgl.weekday()],
            "total_kg": total,
            "unpad_kg": unpad,
            "ipdn_kg": ipdn,
            "by_category_kg": cats,
            "note": str(note).strip() if note else None,
            "quality_flag": str(qf).strip() if qf else None,
        })
    entries.sort(key=lambda e: e["date"])
    return entries


def month_summary(month: str, label: str, dim: int, entries: list[dict]) -> dict:
    active = [e for e in entries if e["total_kg"] > 0 and e["quality_flag"] != "excluded_from_average"]
    days_active = len(active)
    sums = {c: round(sum(e["by_category_kg"][c] for e in entries), 3) for c in CATEGORY_COLS}
    unpad_kg = round(sum(e["unpad_kg"] for e in entries), 3)
    ipdn_kg = round(sum(e["ipdn_kg"] for e in entries), 3)
    total_kg = round(unpad_kg + ipdn_kg, 3)
    return {
        "month": month,
        "label": label,
        "total_kg": total_kg if total_kg else None,
        "organik_anorganik_kg": sums["organik_anorganik"],
        "sisa_makanan_kg": sums["sisa_makanan"],
        "lingkungan_kg": sums["lingkungan"],
        "aset_kg": sums["aset"],
        "avg_kg_per_calendar_day": round(total_kg / dim, 2) if dim and total_kg else None,
        "days_active": days_active,
        "days_in_month": dim,
        "avg_kg_per_active_day": round(total_kg / days_active, 2) if days_active else None,
        "category_breakdown_available": bool(any(sums.values())),
        "unpad_kg": unpad_kg,
        "ipdn_kg": ipdn_kg,
        "ipdn_active_days": sum(1 for e in entries if e["ipdn_kg"] > 0),
    }


def dump_source_csv(out_dir: Path, daily: list[dict]) -> Path:
    """Snapshot teks dari sumber, supaya git punya riwayatnya (xlsx di-gitignore)."""
    path = out_dir / "_sources" / "timbulan_master.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    cols = ["tanggal", "hari", *[f"{c}_kg" for c in CATEGORY_COLS], "ipdn_kg", "catatan", "quality_flag"]
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for e in daily:
            w.writerow([e["date"], e["day_of_week"], *[e["by_category_kg"][c] for c in CATEGORY_COLS],
                        e["ipdn_kg"], e["note"] or "", e["quality_flag"] or ""])
    return path


def main() -> int:
    ap = argparse.ArgumentParser(description=f"[PENSIUN] Build {DATASET_ID}.json dari workbook induk")
    ap.add_argument("--source", default=None, help=f"Path ke {MASTER_NAME}")
    ap.add_argument("--out", default=None)
    ap.add_argument("--i-know-this-is-retired", action="store_true")
    args = ap.parse_args()

    if not args.i_know_this_is_retired:
        print(
            "[parse_timbulan_master] SKRIP INI SUDAH PENSIUN dan menolak jalan.\n"
            "  Sumber resmi sekarang adalah BUKU BESAR: data/_ledger/timbulan.csv\n"
            "  Ia melewati buku besar, sehingga baris yang hanya ada di buku besar bisa hilang.\n"
            "  Pakai:  python parse_timbulan_ledger.py --out data\n"
            "  Tambah data baru:  python import_inbox.py --dataset timbulan",
            file=sys.stderr,
        )
        return 1

    src = locate_master(args.source)
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)

    required = ["Harian", "Bulan", "Kategori", "Kendaraan", "Tare Kontainer", "Catatan Kualitas"]
    missing = [s for s in required if s not in wb.sheetnames]
    if missing:
        raise SystemExit(f"[parse_timbulan_master] sheet hilang di {src.name}: {missing}. ABORT.")

    daily = read_daily(wb["Harian"])
    if not daily:
        raise SystemExit(
            f"[parse_timbulan_master] sheet 'Harian' tidak berisi satu pun hari dengan timbangan.\n"
            f"Menolak menulis timbulan.json kosong. ABORT."
        )

    months = rows(wb["Bulan"])
    monthly = [
        month_summary(str(m["month"]), str(m["label"]), int(m["days_in_month"]),
                      [e for e in daily if e["date"].startswith(str(m["month"]))])
        for m in months
    ]

    data = base_dataset(
        DATASET_ID,
        source_files=[MASTER_NAME],
        period={"start": daily[0]["date"], "end": daily[-1]["date"]},
    )
    data["generated_at"] = dt.date.today().isoformat()
    data["data_quality_flags"] = [
        {"severity": str(f["severity"]), "message": str(f["message"])}
        for f in rows(wb["Catatan Kualitas"])
    ]
    data["unit_default"] = "kg"
    data["vehicle_sources"] = [
        {k: v for k, v in {
            "id": r["id"], "label": r["label"], "operator": r["operator"],
            "tare_kg": r["tare_kg"], "note": r.get("note") or None,
        }.items() if k != "note" or v}
        for r in rows(wb["Kendaraan"])
    ]
    data["container_tare_kg"] = {str(r["kontainer"]): r["tare_kg"] for r in rows(wb["Tare Kontainer"])}
    data["categories"] = [
        {"id": r["id"], "label": r["label"], "color_key": r["color_key"], "source": r["source"]}
        for r in rows(wb["Kategori"])
    ]
    data["monthly_summary"] = monthly
    data["daily_entries"] = daily
    wb.close()

    out_dir = resolve_output(args.out)
    target = out_dir / f"{DATASET_ID}.json"
    write_json(target, data)
    csv_path = dump_source_csv(out_dir, daily)

    total = sum(e["total_kg"] for e in daily)
    active_months = sum(1 for m in monthly if m["total_kg"])
    print(f"[parse_{DATASET_ID}] sumber: {src}")
    print(f"[parse_{DATASET_ID}] wrote {target} ({len(daily)} hari, {active_months} bulan berisi, {total:,.0f} kg)")
    print(f"[parse_{DATASET_ID}] snapshot sumber: {csv_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
