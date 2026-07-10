"""Parse Logbook Limbah B3.xlsx → b3_waste.json.

Source: 12 sheets covering Sep 2024 – 2026.
- 10 monthly report sheets ('Laporan Limbah B3 (Sep 2024)' … '(Des. 2025)')
  with columns: No | Nama Lembaga | Nama Limbah | Volume | Satuan | Kode Limbah | Kategori
- 1 combined 2026 sheet adds: Bulan | Tanggal columns
- 1 wide Logbook (TPS Masuk/Keluar/Sisa) — skipped here, different structure

Output: dashboard-ready aggregations (by month, by lembaga, by kode limbah)
plus full entry list for client-side filtering.
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _utils import (
    base_dataset,
    find_project_root,
    find_source,
    flag,
    parse_number,
    resolve_output,
    round_to,
    write_json,
)
from _ledger import ledger_dir, read_rows

SOURCE_XLSX = "Logbook Limbah B3.xlsx"
DATASET_ID = "b3_waste"

# Kamus kode limbah ber-provenance (buku besar teks, bukan Excel).
# Excel = kotak masuk baca-saja; kode yang belum diisi di Excel diambil dari sini.
# Kunci baris = (tanggal, lembaga, nama_limbah, volume, satuan) — stabil terhadap
# penyisipan baris Excel, dan mencakup `tanggal` agar dua entri identik pada tanggal
# berbeda (mis. dugaan duplikat 4 Juni vs 8 Juni) tidak saling menimpa.
KAMUS_CSV = "b3_waste_kode.csv"
KAMUS_ALIAS_CSV = "b3_waste_kode_alias.csv"

# --- Satuan -----------------------------------------------------------------
# Satuan dinormalkan SEKALI di sini. Setiap entri terbit dengan volume_liter dan
# mass_kg yang sudah jadi, supaya frontend tidak perlu menebak dari teks satuan.
# Sebelumnya pengecekan `satuan.startswith("liter")` membuang 'Mili Liter'
# (15 entri, 2.485 L) tanpa peringatan apa pun.
SATUAN_KE_LITER = {
    "liter": 1.0, "l": 1.0, "ltr": 1.0, "lt": 1.0,
    "mili liter": 0.001, "mililiter": 0.001, "milliliter": 0.001, "ml": 0.001, "cc": 0.001,
}
SATUAN_KE_KG = {
    "kg": 1.0, "kilogram": 1.0, "kgs": 1.0,
    "gram": 0.001, "gr": 0.001, "g": 0.001,
    "ton": 1000.0,
}
# Satuan cacah: sah, tetapi tidak dapat dijumlahkan ke volume/massa.
SATUAN_CACAH = {"buah", "pcs", "pieces", "unit", "botol", "dus", "drum", "jerigen"}


def klasifikasi_satuan(satuan: str) -> tuple[str, float]:
    """'Mili Liter' -> ('liter', 0.001). Satuan tak dikenal -> ('?', 0.0)."""
    s = " ".join(str(satuan or "").strip().lower().split())
    if not s:
        return ("?", 0.0)
    if s in SATUAN_KE_LITER:
        return ("liter", SATUAN_KE_LITER[s])
    if s in SATUAN_KE_KG:
        return ("kg", SATUAN_KE_KG[s])
    if s in SATUAN_CACAH:
        return ("cacah", 0.0)
    return ("?", 0.0)

# Normalisasi nama lembaga — alias resmi yang dipakai institusi sumber
LEMBAGA_ALIAS = {
    "pusat unggulan ilmu lingkungan": "CESS",
    "puil": "CESS",
}


def normalize_lembaga(name: str) -> str:
    if not name:
        return name
    key = name.strip().lower()
    return LEMBAGA_ALIAS.get(key, name.strip())

MONTHS_ID = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]
MONTH_ALIAS_TO_NUM = {}
for i, m in enumerate(MONTHS_ID, start=1):
    MONTH_ALIAS_TO_NUM[m.lower()] = i
    MONTH_ALIAS_TO_NUM[m[:3].lower()] = i
MONTH_ALIAS_TO_NUM.update({"sept": 9, "okt": 10, "des": 12, "agu": 8, "jun": 6, "mei": 5, "feb": 2})


def month_from_sheet(name: str) -> str | None:
    """Extract YYYY-MM from sheet name like 'Laporan Limbah B3 (Sep 2024)'."""
    m = re.search(r"\(([\w\.\s]+)\s+(\d{4})\)", name)
    if not m:
        return None
    token, year = m.group(1).strip().rstrip("."), int(m.group(2))
    num = MONTH_ALIAS_TO_NUM.get(token.lower())
    if num is None:
        return None
    return f"{year:04d}-{num:02d}"


def month_label(iso: str) -> str:
    """2024-09 → 'September 2024'."""
    try:
        y, m = iso.split("-")
        return f"{MONTHS_ID[int(m) - 1]} {y}"
    except (ValueError, IndexError):
        return iso


def parse_month_from_string(s: str, default_year: int) -> str | None:
    """Parse 'Januari' → '{year}-01'."""
    if not s:
        return None
    num = MONTH_ALIAS_TO_NUM.get(s.strip().lower())
    if num is None:
        return None
    return f"{default_year:04d}-{num:02d}"


def fmt_date(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip() or None


def parse_monthly_sheet(ws, month_iso: str) -> list[dict]:
    """Parse standard 7-col monthly sheet → entries list."""
    entries: list[dict] = []
    # Find header row (has 'No' in col 1)
    header_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        v0 = ws.cell(row=r, column=1).value
        if v0 and str(v0).strip().lower() in ("no", "no."):
            header_row = r
            break
    if header_row is None:
        return entries
    for r in range(header_row + 1, ws.max_row + 1):
        lembaga = ws.cell(row=r, column=2).value
        limbah = ws.cell(row=r, column=3).value
        if not lembaga or not limbah:
            continue
        volume = parse_number(ws.cell(row=r, column=4).value)
        satuan = ws.cell(row=r, column=5).value
        kode = ws.cell(row=r, column=6).value
        kategori = ws.cell(row=r, column=7).value
        if volume is None:
            continue
        entries.append({
            "month": month_iso,
            "date": None,
            "lembaga": normalize_lembaga(str(lembaga)),
            "limbah": str(limbah).strip(),
            "volume": round_to(float(volume), 3),
            "satuan": (str(satuan).strip() if satuan else ""),
            "kode_limbah": (str(kode).strip() if kode else ""),
            "kategori": (str(kategori).strip() if kategori else ""),
        })
    return entries


def parse_2026_sheet(ws) -> list[dict]:
    """Parse the combined 2026 sheet — has Bulan + Tanggal columns."""
    entries: list[dict] = []
    # Header row 1
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip().lower()] = c
    col_bulan = headers.get("bulan")
    col_tanggal = headers.get("tanggal")
    col_lembaga = headers.get("nama lembaga")
    col_limbah = headers.get("nama limbah")
    col_volume = headers.get("volume")
    col_satuan = headers.get("satuan")
    col_kode = headers.get("kode limbah")
    col_kategori = headers.get("kategori")
    if not col_lembaga or not col_volume:
        return entries
    for r in range(2, ws.max_row + 1):
        lembaga = ws.cell(row=r, column=col_lembaga).value
        if not lembaga:
            continue
        bulan_raw = ws.cell(row=r, column=col_bulan).value if col_bulan else None
        tanggal_raw = ws.cell(row=r, column=col_tanggal).value if col_tanggal else None
        iso_date = fmt_date(tanggal_raw)
        month_iso = None
        if iso_date and re.match(r"^\d{4}-\d{2}-\d{2}$", iso_date):
            month_iso = iso_date[:7]
        elif bulan_raw:
            month_iso = parse_month_from_string(str(bulan_raw), default_year=2026)
        if not month_iso:
            continue
        volume = parse_number(ws.cell(row=r, column=col_volume).value)
        if volume is None:
            continue
        entries.append({
            "month": month_iso,
            "date": iso_date,
            "lembaga": normalize_lembaga(str(lembaga)),
            "limbah": str(ws.cell(row=r, column=col_limbah).value or "").strip(),
            "volume": round_to(float(volume), 3),
            "satuan": str(ws.cell(row=r, column=col_satuan).value or "").strip(),
            "kode_limbah": str(ws.cell(row=r, column=col_kode).value or "").strip(),
            "kategori": str(ws.cell(row=r, column=col_kategori).value or "").strip(),
        })
    return entries


def lengkapi_satuan(entries: list[dict]) -> tuple[list[str], int]:
    """Isi volume_liter + mass_kg tiap entri. Kembalikan (satuan_tak_dikenal, jml_cacah)."""
    tak_dikenal: dict[str, int] = {}
    cacah = 0
    for e in entries:
        jenis, faktor = klasifikasi_satuan(e["satuan"])
        vol = float(e["volume"])
        e["volume_liter"] = round_to(vol * faktor, 4) if jenis == "liter" else 0.0
        e["mass_kg"] = round_to(vol * faktor, 4) if jenis == "kg" else 0.0
        if jenis == "cacah":
            cacah += 1
        elif jenis == "?":
            key = str(e["satuan"] or "(kosong)")
            tak_dikenal[key] = tak_dikenal.get(key, 0) + 1
    rincian = [f"'{k}' ({n} entri)" for k, n in sorted(tak_dikenal.items())]
    return rincian, cacah


# --- Logbook TPS: limbah masuk, keluar, dan sisa -----------------------------
# Sheet `Logbook (Sep 24 - Jan 26)` memuat DUA tabel berdampingan dalam satu grid:
#   kolom 1-6   Limbah Masuk   : No | Tanggal Masuk | Kode | Sumber | Kg | Maksimal Penyimpanan
#   kolom 8-13  Limbah Keluar  : Sumber | Tanggal Keluar | Kode | Kg | Tujuan | Bukti Dokumen
# Baris keduanya TIDAK sejajar. Tanggal & tujuan keluar hanya ditulis di baris pertama
# tiap pengiriman, jadi harus dibawa turun (forward-fill).
#
# Batas antar-pengiriman ditandai baris berlabel "Total Jumlah ..." atau
# "Total Limbah yang keluar ...". Kami memakai label itu, BUKAN nomor baris, supaya
# tidak rusak bila pemilik menyisipkan baris.
LOGBOOK_SHEET_HINT = "Logbook"
BATAS_PENGIRIMAN = ("total jumlah", "total limbah yang keluar")


def _teks(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    return str(v).strip()


def _tanggal_logbook(v) -> tuple[str, bool]:
    """-> (iso, benar_tanggal). '30/06/2025' disimpan sebagai TEKS di sumber."""
    if isinstance(v, datetime):
        return v.date().isoformat(), True
    if isinstance(v, date):
        return v.isoformat(), True
    t = _teks(v)
    if not t:
        return "", True
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", t)
    if m:
        d, mo, y = (int(x) for x in m.groups())
        return f"{y:04d}-{mo:02d}-{d:02d}", False
    return t, False


def parse_tps_logbook(ws) -> dict:
    """Baca sheet logbook TPS -> {masuk, keluar, pengiriman, catatan}."""
    # Cari baris header (punya 'Tanggal Masuk' di suatu kolom).
    hdr = None
    for r in range(1, min(ws.max_row, 12) + 1):
        labels = [_teks(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)]
        if "tanggal masuk" in labels and "tanggal keluar" in labels:
            hdr = r
            break
    if hdr is None:
        return {}

    kol = {}
    for c in range(1, ws.max_column + 1):
        lab = _teks(ws.cell(hdr, c).value).lower()
        if lab and lab not in kol:
            kol[lab] = c
    c_tgl_masuk = kol["tanggal masuk"]
    c_tgl_keluar = kol["tanggal keluar"]
    c_kode_masuk = c_tgl_masuk + 1
    c_sumber = c_tgl_masuk + 2
    c_kg_masuk = c_tgl_masuk + 3
    c_maks = c_tgl_masuk + 4
    c_kode_keluar = c_tgl_keluar + 1
    c_kg_keluar = c_tgl_keluar + 2
    c_tujuan = c_tgl_keluar + 3
    c_bukti = c_tgl_keluar + 4

    masuk: list[dict] = []
    keluar: list[dict] = []
    catatan: list[str] = []
    batas_masuk: list[int] = []     # indeks pemisah batch di daftar `masuk`
    tgl_keluar_kini = ""
    tujuan_kini = ""
    bukti_kini = ""
    tanggal_teks: list[str] = []

    for r in range(hdr + 1, ws.max_row + 1):
        label = _teks(ws.cell(r, 1).value).lower()
        if label.startswith(BATAS_PENGIRIMAN):
            batas_masuk.append(len(masuk))

        # --- sisi masuk
        tgl_raw = ws.cell(r, c_tgl_masuk).value
        kode = _teks(ws.cell(r, c_kode_masuk).value)
        kg = ws.cell(r, c_kg_masuk).value
        if tgl_raw is not None and kode and kg is not None:
            iso, ok = _tanggal_logbook(tgl_raw)
            if not ok:
                tanggal_teks.append(f"masuk baris {r}: {_teks(tgl_raw)!r}")
            masuk.append({
                "tanggal": iso,
                "kode_limbah": normalkan_kode(kode),
                "sumber": normalize_lembaga(_teks(ws.cell(r, c_sumber).value)),
                "mass_kg": round_to(float(kg), 3),
                "maksimal_penyimpanan": _teks(ws.cell(r, c_maks).value) or None,
            })

        # --- sisi keluar (tanggal/tujuan/bukti dibawa turun)
        tgl_k_raw = ws.cell(r, c_tgl_keluar).value
        if tgl_k_raw is not None and _teks(tgl_k_raw):
            iso, ok = _tanggal_logbook(tgl_k_raw)
            if not ok:
                tanggal_teks.append(f"keluar baris {r}: {_teks(tgl_k_raw)!r}")
            tgl_keluar_kini = iso
            tujuan_kini = _teks(ws.cell(r, c_tujuan).value) or tujuan_kini
            bukti_kini = _teks(ws.cell(r, c_bukti).value) or bukti_kini
        kode_k = _teks(ws.cell(r, c_kode_keluar).value)
        kg_k = ws.cell(r, c_kg_keluar).value
        if kode_k and kg_k is not None:
            keluar.append({
                "tanggal": tgl_keluar_kini,
                "kode_limbah": normalkan_kode(kode_k),
                "mass_kg": round_to(float(kg_k), 3),
                "tujuan": tujuan_kini or None,
                "bukti_dokumen": bukti_kini or None,
            })

    if tanggal_teks:
        catatan.append(
            "Tanggal disimpan sebagai teks, bukan tipe tanggal Excel: "
            + "; ".join(tanggal_teks)
        )

    # --- kelompokkan menjadi pengiriman
    pengiriman = _bentuk_pengiriman(masuk, keluar, batas_masuk, catatan)
    return {"masuk": masuk, "keluar": keluar, "pengiriman": pengiriman, "catatan": catatan}


def _bentuk_pengiriman(masuk, keluar, batas_masuk, catatan) -> list[dict]:
    """Pasangkan tiap tanggal keluar dengan blok limbah masuk yang mendahuluinya."""
    from collections import defaultdict

    urut_tgl: list[str] = []
    for k in keluar:
        if k["tanggal"] and k["tanggal"] not in urut_tgl:
            urut_tgl.append(k["tanggal"])

    potong = [0] + batas_masuk
    hasil = []
    for i, tgl in enumerate(urut_tgl):
        blok = masuk[potong[i]:potong[i + 1]] if i + 1 < len(potong) else masuk[potong[i]:]
        keluar_i = [k for k in keluar if k["tanggal"] == tgl]

        m_kode = defaultdict(float)
        for m in blok:
            m_kode[m["kode_limbah"]] += m["mass_kg"]
        k_kode = defaultdict(float)
        for k in keluar_i:
            k_kode[k["kode_limbah"]] += k["mass_kg"]

        tgl_blok = sorted(m["tanggal"] for m in blok if m["tanggal"])
        total_m = round_to(sum(m_kode.values()), 3)
        total_k = round_to(sum(k_kode.values()), 3)

        selisih = [
            {"kode_limbah": kd,
             "masuk_kg": round_to(m_kode.get(kd, 0.0), 3),
             "keluar_kg": round_to(k_kode.get(kd, 0.0), 3),
             "selisih_kg": round_to(m_kode.get(kd, 0.0) - k_kode.get(kd, 0.0), 3)}
            for kd in sorted(set(m_kode) | set(k_kode))
            if abs(m_kode.get(kd, 0.0) - k_kode.get(kd, 0.0)) > 1e-6
        ]
        for s in selisih:
            catatan.append(
                f"Pengiriman {tgl}: kode {s['kode_limbah']} masuk {s['masuk_kg']} kg "
                f"tetapi keluar {s['keluar_kg']} kg (selisih {s['selisih_kg']} kg)."
            )
        janggal = [m for m in blok if m["tanggal"] and tgl and m["tanggal"] > tgl]
        if janggal:
            catatan.append(
                f"Pengiriman {tgl}: {len(janggal)} baris limbah masuk bertanggal SETELAH "
                f"tanggal keluar ({', '.join(sorted({m['tanggal'] for m in janggal}))}). "
                f"Mustahil terangkut; kemungkinan salah ketik bulan di sumber."
            )

        hasil.append({
            "tanggal_keluar": tgl,
            "tujuan": next((k["tujuan"] for k in keluar_i if k["tujuan"]), None),
            "bukti_dokumen": next((k["bukti_dokumen"] for k in keluar_i if k["bukti_dokumen"]), None),
            "periode_masuk": {"start": tgl_blok[0] if tgl_blok else None,
                              "end": tgl_blok[-1] if tgl_blok else None},
            "entri_masuk": len(blok),
            "total_masuk_kg": total_m,
            "total_keluar_kg": total_k,
            "selisih_kg": round_to(total_m - total_k, 3),
            "selisih_per_kode": selisih,
        })
    return hasil


def rangkum_tps(tps: dict) -> dict:
    """Ringkas logbook TPS + hitung sisa per kode (masuk - keluar)."""
    from collections import defaultdict

    m_kode = defaultdict(float)
    for m in tps["masuk"]:
        m_kode[m["kode_limbah"]] += m["mass_kg"]
    k_kode = defaultdict(float)
    for k in tps["keluar"]:
        k_kode[k["kode_limbah"]] += k["mass_kg"]

    sisa = [
        {"kode_limbah": kd,
         "masuk_kg": round_to(m_kode.get(kd, 0.0), 3),
         "keluar_kg": round_to(k_kode.get(kd, 0.0), 3),
         "sisa_kg": round_to(m_kode.get(kd, 0.0) - k_kode.get(kd, 0.0), 3)}
        for kd in sorted(set(m_kode) | set(k_kode))
    ]
    total_m = round_to(sum(m_kode.values()), 3)
    total_k = round_to(sum(k_kode.values()), 3)
    tgl_masuk = sorted(m["tanggal"] for m in tps["masuk"] if m["tanggal"])

    return {
        "summary": {
            "entri_masuk": len(tps["masuk"]),
            "entri_keluar": len(tps["keluar"]),
            "pengiriman": len(tps["pengiriman"]),
            "total_masuk_kg": total_m,
            "total_keluar_kg": total_k,
            "sisa_di_tps_kg": round_to(total_m - total_k, 3),
            "periode_masuk": {"start": tgl_masuk[0] if tgl_masuk else None,
                              "end": tgl_masuk[-1] if tgl_masuk else None},
        },
        "pengiriman": tps["pengiriman"],
        "sisa_per_kode": [s for s in sisa if abs(s["sisa_kg"]) > 1e-6],
        "masuk": tps["masuk"],
        "keluar": tps["keluar"],
    }


# --- Kamus kode limbah ------------------------------------------------------

def _canon_vol(v) -> str:
    """Kanonkan volume untuk kunci: 1.0 -> '1', 0.5 -> '0.5'. Sama di CSV & entri."""
    r = round(float(v), 3)
    return str(int(r)) if r == int(r) else str(r)


def kamus_key(tanggal, lembaga, nama, volume, satuan) -> tuple:
    """Kunci gabungan stabil. Nama & lembaga di-strip; satuan di-lowercase."""
    return (
        str(tanggal or "").strip(),
        str(lembaga or "").strip(),
        str(nama or "").strip(),
        _canon_vol(volume),
        " ".join(str(satuan or "").strip().lower().split()),
    )


def load_kamus(led: Path) -> dict[tuple, dict]:
    """Muat b3_waste_kode.csv menjadi {kunci: baris}. Kosong bila file tak ada."""
    rows = read_rows(led / KAMUS_CSV)
    kamus: dict[tuple, dict] = {}
    for r in rows:
        k = kamus_key(r.get("tanggal"), r.get("lembaga"), r.get("nama_limbah"),
                      r.get("volume"), r.get("satuan"))
        if k in kamus:
            # Kamus wajib unik per kunci; tabrakan = kesalahan data yang harus terlihat.
            raise SystemExit(f"[parse_b3_waste] GAGAL: kunci kamus ganda di {KAMUS_CSV}: {k}")
        kamus[k] = r
    return kamus


def load_alias(led: Path) -> dict[str, dict]:
    """Muat b3_waste_kode_alias.csv menjadi {dari: baris}. Untuk flip industri 38->37
    tanpa menyentuh kode: cukup tambah satu baris CSV di sini."""
    alias: dict[str, dict] = {}
    for r in read_rows(led / KAMUS_ALIAS_CSV):
        dari = (r.get("dari") or "").strip()
        if dari:
            alias[dari] = r
    return alias


# --- Notasi kode limbah -----------------------------------------------------
# Lampiran IX menulis kode Tabel 1 sebagai huruf besar + tiga angka + sufiks huruf
# KECIL (A106d, B104d), dan kode industri sebagai huruf besar + angka + '-' + angka
# (A337-1). Logbook menulis sufiksnya huruf besar (A106D). Bila keduanya dibiarkan,
# `by_kode_limbah` memecah A106D (151 entri) dari A106d (25 entri) — satu kode yang
# sama tampil sebagai dua kotak di treemap. Provenance kode TIDAK boleh dititipkan
# pada besar-kecil huruf; itu tugas `kode_status`.
KODE_TABEL1 = re.compile(r"^([AB])(\d{3})([a-dA-D])$")
KODE_INDUSTRI = re.compile(r"^([AB])(\d{3})-(\d+)$")


def normalkan_kode(kode: str) -> str:
    """'A106D' -> 'A106d'. Bentuk yang tak dikenali dikembalikan apa adanya."""
    k = str(kode or "").strip()
    if not k:
        return ""
    m = KODE_TABEL1.match(k)
    if m:
        return f"{m.group(1).upper()}{m.group(2)}{m.group(3).lower()}"
    m = KODE_INDUSTRI.match(k)
    if m:
        return f"{m.group(1).upper()}{m.group(2)}-{m.group(3)}"
    return k


def terapkan_kamus(entries: list[dict], kamus: dict[tuple, dict],
                   alias: dict[str, dict]) -> dict:
    """Isi kode_limbah dari kamus HANYA bila sel Excel kosong. Excel selalu menang.

    Menetapkan `kode_sumber` (excel/kamus/kosong) dan `kode_status` pada setiap entri.
    Mengembalikan statistik + daftar konflik & alias untuk diterbitkan sebagai flag.
    """
    n_kamus = n_excel = n_kosong = 0
    konflik: list[str] = []
    alias_terpakai: set[tuple] = set()
    for e in entries:
        k = kamus_key(e["date"], e["lembaga"], e["limbah"], e["volume"], e["satuan"])
        match = kamus.get(k)
        if e["kode_limbah"]:
            # Excel sudah mengisi kode — Excel menang. Bila kamus beda, jangan diam.
            e["kode_sumber"] = "excel"
            e["kode_status"] = "tercatat"
            n_excel += 1
            if match and match.get("kode") and \
               match["kode"].strip().lower() != e["kode_limbah"].strip().lower():
                konflik.append(
                    f"{e['date']} {e['lembaga']} '{e['limbah'][:40]}': "
                    f"Excel='{e['kode_limbah']}' vs kamus='{match['kode']}' "
                    f"(status {match.get('status')}). Excel dipakai."
                )
        elif match:
            kode = (match.get("kode") or "").strip()
            a = alias.get(kode)
            if a and (a.get("ke") or "").strip():
                alias_terpakai.add((kode, a["ke"].strip()))
                kode = a["ke"].strip()
            e["kode_limbah"] = kode
            e["kode_sumber"] = "kamus"
            e["kode_status"] = (match.get("status") or "usulan").strip()
            n_kamus += 1
        else:
            # Tetap tanpa kode: masuk keranjang '—', flag tetap terbit.
            e["kode_sumber"] = "kosong"
            e["kode_status"] = ""
            n_kosong += 1

    # Selaraskan notasi SETELAH sumber kode ditetapkan, berlaku sama bagi kode dari
    # Excel maupun dari kamus. Yang bentuknya asing tidak diubah, tetapi dilaporkan.
    n_normal = 0
    bentuk_asing: set[str] = set()
    for e in entries:
        asli = e["kode_limbah"]
        if not asli:
            continue
        baru = normalkan_kode(asli)
        if baru != asli:
            n_normal += 1
            e["kode_limbah"] = baru
        if not (KODE_TABEL1.match(baru) or KODE_INDUSTRI.match(baru)):
            bentuk_asing.add(baru)

    return {
        "n_kamus": n_kamus,
        "n_excel": n_excel,
        "n_kosong": n_kosong,
        "konflik": konflik,
        "alias_terpakai": sorted(alias_terpakai),
        "n_normal": n_normal,
        "bentuk_asing": sorted(bentuk_asing),
        # Sebaran status kode yang berasal dari kamus. Dipakai untuk menentukan
        # tingkat flag: 'usulan' = warning, 'disahkan' = info.
        "status_kamus": Counter(
            e["kode_status"] for e in entries if e["kode_sumber"] == "kamus"
        ),
    }


def deteksi_duplikat(entries: list[dict]) -> list[str]:
    """Dugaan duplikat: (lembaga, limbah, volume, satuan) sama, bulan sama, tanggal beda.

    Tidak menghapus apa pun — hanya menerbitkan peringatan agar 2 Kg yang mungkin
    terhitung dua kali (baris 205/206 vs 217/203) terlihat oleh peninjau.
    """
    from collections import defaultdict
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for e in entries:
        if not e["date"]:
            continue
        groups[(e["lembaga"], e["limbah"], _canon_vol(e["volume"]), e["satuan"])].append(e)
    pesan: list[str] = []
    for (lembaga, limbah, vol, satuan), grp in groups.items():
        tanggal = sorted({e["date"] for e in grp})
        bulan = {e["month"] for e in grp}
        if len(tanggal) > 1 and len(bulan) == 1:
            pesan.append(
                f"Dugaan duplikat: {lembaga} '{limbah[:45]}' {vol} {satuan} "
                f"tercatat pada {', '.join(tanggal)} (bulan {next(iter(bulan))}). "
                f"Bila salin-tempel, {vol} {satuan} terhitung dua kali. "
                f"Tidak dihapus — perlu konfirmasi pemilik."
            )
    return sorted(pesan)


def aggregate(entries: list[dict]) -> dict:
    """Compute summary + per-month + per-lembaga + per-kode aggregations."""
    from collections import defaultdict

    # Per month
    by_month: dict[str, dict] = defaultdict(lambda: {
        "entries": 0, "volume_liter": 0.0, "mass_kg": 0.0,
        "by_kategori": defaultdict(lambda: {"volume_liter": 0.0, "mass_kg": 0.0}),
    })
    # Per lembaga
    by_lembaga: dict[str, dict] = defaultdict(lambda: {
        "entries": 0, "volume_liter": 0.0, "mass_kg": 0.0,
    })
    # Per kode limbah
    by_kode: dict[str, dict] = defaultdict(lambda: {
        "entries": 0, "volume_liter": 0.0, "mass_kg": 0.0, "kategori": set(),
    })

    total_liter = total_kg = 0.0
    for e in entries:
        kat = (e["kategori"] or "").lower() or "unknown"
        m_key = e["month"]
        l_key = e["lembaga"]
        k_key = e["kode_limbah"] or "—"

        # volume_liter / mass_kg sudah dinormalkan oleh lengkapi_satuan().
        # Entri bersatuan cacah (Buah) bernilai 0 pada keduanya — tetap terhitung
        # sebagai entri, tetapi tidak menambah volume maupun massa.
        liter = float(e["volume_liter"])
        kg = float(e["mass_kg"])

        by_month[m_key]["entries"] += 1
        by_lembaga[l_key]["entries"] += 1
        by_kode[k_key]["entries"] += 1
        if e["kategori"]:
            by_kode[k_key]["kategori"].add(e["kategori"])

        if liter:
            by_month[m_key]["volume_liter"] += liter
            by_month[m_key]["by_kategori"][kat]["volume_liter"] += liter
            by_lembaga[l_key]["volume_liter"] += liter
            by_kode[k_key]["volume_liter"] += liter
            total_liter += liter
        if kg:
            by_month[m_key]["mass_kg"] += kg
            by_month[m_key]["by_kategori"][kat]["mass_kg"] += kg
            by_lembaga[l_key]["mass_kg"] += kg
            by_kode[k_key]["mass_kg"] += kg
            total_kg += kg

    def freeze_kategori_breakdown(d):
        return {k: {"volume_liter": round_to(v["volume_liter"], 2),
                    "mass_kg": round_to(v["mass_kg"], 2)} for k, v in d.items()}

    monthly_totals = [
        {
            "month": m,
            "label": month_label(m),
            "entries": d["entries"],
            "volume_liter": round_to(d["volume_liter"], 2),
            "mass_kg": round_to(d["mass_kg"], 2),
            "by_kategori": freeze_kategori_breakdown(d["by_kategori"]),
        }
        for m, d in sorted(by_month.items())
    ]
    lembaga_totals = sorted([
        {
            "lembaga": l,
            "entries": d["entries"],
            "volume_liter": round_to(d["volume_liter"], 2),
            "mass_kg": round_to(d["mass_kg"], 2),
        }
        for l, d in by_lembaga.items()
    ], key=lambda x: -(x["volume_liter"] + x["mass_kg"]))

    kode_totals = sorted([
        {
            "kode": k,
            "entries": d["entries"],
            "volume_liter": round_to(d["volume_liter"], 2),
            "mass_kg": round_to(d["mass_kg"], 2),
            "kategori": sorted(d["kategori"]),
        }
        for k, d in by_kode.items()
    ], key=lambda x: -(x["volume_liter"] + x["mass_kg"]))

    return {
        "summary": {
            "total_entries": len(entries),
            "total_volume_liter": round_to(total_liter, 2),
            "total_mass_kg": round_to(total_kg, 2),
            "unique_lembaga": len(by_lembaga),
            "unique_kode_limbah": len(by_kode),
            "months_with_data": len(by_month),
        },
        "monthly_totals": monthly_totals,
        "by_lembaga": lembaga_totals,
        "by_kode_limbah": kode_totals,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=f"Build {DATASET_ID}.json")
    ap.add_argument("--source", default=None)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("[parse_b3_waste] openpyxl not installed; run: pip install openpyxl", file=sys.stderr)
        return 1

    project = find_project_root()
    if args.source:
        src = Path(args.source)
    else:
        # *.xlsx di-gitignore, jadi sumbernya tinggal di folder data user, bukan root repo.
        src = find_source(SOURCE_XLSX, subdirs=("Limbah B3",)) or (project / SOURCE_XLSX)
    if not src.exists():
        print(f"[parse_b3_waste] source not found: {src}", file=sys.stderr)
        print(f"[parse_b3_waste] dicari juga di '<Dashboard>/Data dan Pengetahuan/Limbah B3/'.", file=sys.stderr)
        return 1
    print(f"[parse_b3_waste] sumber: {src}")

    wb = openpyxl.load_workbook(src, data_only=True)

    entries: list[dict] = []
    skipped_sheets: list[str] = []
    tps: dict = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Monthly report sheet?
        month_iso = month_from_sheet(sheet_name)
        if month_iso:
            entries.extend(parse_monthly_sheet(ws, month_iso))
            continue
        # 2026 combined sheet?
        if "2026" in sheet_name and "Laporan" in sheet_name:
            entries.extend(parse_2026_sheet(ws))
            continue
        # Sheet logbook TPS (Limbah Masuk / Keluar / Sisa)
        if LOGBOOK_SHEET_HINT in sheet_name:
            tps = parse_tps_logbook(ws)
            if not tps:
                print(
                    f"[parse_b3_waste] GAGAL: sheet '{sheet_name}' dikenali sebagai logbook TPS "
                    f"tetapi header 'Tanggal Masuk'/'Tanggal Keluar' tidak ditemukan.",
                    file=sys.stderr,
                )
                return 1
            continue
        skipped_sheets.append(sheet_name)

    if not entries:
        print("[parse_b3_waste] no entries parsed", file=sys.stderr)
        return 1

    # Normalkan satuan sebelum agregasi. Satuan tak dikenal = GAGAL-KERAS:
    # lebih baik build berhenti daripada volume hilang diam-diam dari total.
    satuan_tak_dikenal, jml_cacah = lengkapi_satuan(entries)
    if satuan_tak_dikenal:
        print(
            "[parse_b3_waste] GAGAL: satuan tidak dikenal — "
            + ", ".join(satuan_tak_dikenal),
            file=sys.stderr,
        )
        print(
            "[parse_b3_waste] Daftarkan satuan itu di SATUAN_KE_LITER / SATUAN_KE_KG / "
            "SATUAN_CACAH, jangan biarkan terbuang dari total.",
            file=sys.stderr,
        )
        return 1

    # Isi kode limbah dari kamus repo (hanya bila sel Excel kosong). Setiap entri
    # dapat kode_sumber + kode_status. Excel selalu menang atas kamus.
    led = ledger_dir()
    kamus = load_kamus(led)
    alias = load_alias(led)
    kode_stat = terapkan_kamus(entries, kamus, alias)
    dup_pesan = deteksi_duplikat(entries)

    aggregated = aggregate(entries)

    # Period from entries
    months = sorted({e["month"] for e in entries})
    start = f"{months[0]}-01"
    # End: last day of last month (simple approximation)
    last_y, last_m = months[-1].split("-")
    last_dates = [e["date"] for e in entries if e["date"] and e["month"] == months[-1]]
    end = max(last_dates) if last_dates else f"{months[-1]}-28"

    flags = []
    if skipped_sheets:
        flags.append(flag(
            "info",
            f"Sheet tidak dikenali dan tidak diolah: {', '.join(skipped_sheets)}.",
        ))
    for c in tps.get("catatan", []):
        flags.append(flag("warning", f"Logbook TPS — {c}"))
    if jml_cacah:
        flags.append(flag(
            "warning",
            f"{jml_cacah} entri bersatuan cacah (Buah/pcs) terhitung sebagai entri "
            f"tetapi tidak menambah total volume maupun massa.",
        ))
    # Kode dari kamus = USULAN, belum disahkan PK3L. Jangan disamarkan jadi resmi.
    # Tingkat flag mengikuti status kode di kamus, bukan ditetapkan di muka.
    # 'usulan' belum sah -> warning. 'disahkan' sudah sah -> info.
    n_usulan = kode_stat["status_kamus"].get("usulan", 0)
    n_disahkan = kode_stat["status_kamus"].get("disahkan", 0)
    if n_usulan:
        flags.append(flag(
            "warning",
            f"{n_usulan} entri memakai kode limbah berstatus USULAN dari kamus repo "
            f"({KAMUS_CSV}) karena sel Kode Limbah di Excel kosong. Kode ini belum disahkan "
            f"penanggung jawab limbah B3 UNPAD dan tidak boleh dipakai untuk manifest "
            f"atau pelaporan sebelum disahkan.",
        ))
    if n_disahkan:
        flags.append(flag(
            "info",
            f"{n_disahkan} entri memakai kode limbah dari kamus repo ({KAMUS_CSV}) karena "
            f"sel Kode Limbah di Excel kosong. Kode berstatus DISAHKAN; tanggal dan pihak "
            f"yang mengesahkan tercatat di kamus.",
        ))
    lain = set(kode_stat["status_kamus"]) - {"usulan", "disahkan"}
    if lain:
        flags.append(flag(
            "warning",
            f"Status kode kamus di luar 'usulan'/'disahkan': {sorted(lain)}. Periksa {KAMUS_CSV}.",
        ))
    if kode_stat["n_kosong"]:
        flags.append(flag(
            "warning",
            f"{kode_stat['n_kosong']} entri tetap tanpa kode limbah (tidak ada di kamus) "
            f"dan dikelompokkan di bawah '—'. Isi kolom 'Kode Limbah' di {SOURCE_XLSX} "
            f"atau tambahkan ke {KAMUS_CSV}.",
        ))
    for c in kode_stat["konflik"]:
        flags.append(flag(
            "warning",
            f"Kode Excel berbeda dari usulan kamus — {c}",
        ))
    if kode_stat["alias_terpakai"]:
        flags.append(flag(
            "info",
            "Alias kode kamus diterapkan: "
            + ", ".join(f"{a}->{b}" for a, b in kode_stat["alias_terpakai"])
            + f" (dari {KAMUS_ALIAS_CSV}).",
        ))
    if kode_stat["n_normal"]:
        flags.append(flag(
            "info",
            f"{kode_stat['n_normal']} kode diselaraskan ke notasi Lampiran IX "
            f"(sufiks huruf kecil, mis. 'A106D' ditulis 'A106d'). Penyelarasan ini murni "
            f"penulisan: tanpa itu satu kode yang sama terpecah dua di agregasi per kode.",
        ))
    if kode_stat["bentuk_asing"]:
        flags.append(flag(
            "warning",
            "Kode dengan bentuk di luar pola Lampiran IX (Axxxy / Bxxxy / Axxx-n): "
            + ", ".join(kode_stat["bentuk_asing"])
            + ". Periksa ejaannya di sumber.",
        ))
    for p in dup_pesan:
        flags.append(flag("warning", p))

    data = base_dataset(
        DATASET_ID,
        source_files=[SOURCE_XLSX],
        period={"start": start, "end": end},
    )
    data["data_quality_flags"] = flags
    data.update(aggregated)
    data["entries"] = entries
    if tps:
        data["tps_logbook"] = rangkum_tps(tps)

    out_dir = resolve_output(args.out)
    target = out_dir / f"{DATASET_ID}.json"
    write_json(target, data)
    s = aggregated["summary"]
    print(
        f"[parse_b3_waste] wrote {target} "
        f"({s['total_entries']} entries, "
        f"{s['total_volume_liter']} L + {s['total_mass_kg']} kg, "
        f"{s['unique_lembaga']} lembaga, {s['unique_kode_limbah']} kode, "
        f"{s['months_with_data']} bulan, {len(flags)} flags)"
    )
    if tps:
        t = data["tps_logbook"]["summary"]
        print(
            f"[parse_b3_waste] logbook TPS: {t['total_masuk_kg']} kg masuk, "
            f"{t['total_keluar_kg']} kg keluar dalam {t['pengiriman']} pengiriman, "
            f"sisa {t['sisa_di_tps_kg']} kg."
        )
    rincian_kamus = ", ".join(
        f"{n} {st}" for st, n in sorted(kode_stat["status_kamus"].items())
    ) or "-"
    print(
        f"[parse_b3_waste] kode: {kode_stat['n_excel']} dari Excel, "
        f"{kode_stat['n_kamus']} dari kamus ({rincian_kamus}), "
        f"{kode_stat['n_kosong']} tetap kosong; "
        f"{len(kode_stat['konflik'])} konflik Excel-vs-kamus, {len(dup_pesan)} dugaan duplikat."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
