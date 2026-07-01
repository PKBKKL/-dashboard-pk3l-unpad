"""Rebuild data/timbulan.json with the 4-category vehicle-based schema.

Reads the Excel workbook directly and applies the confirmed business rules:

1. Sampah Organik dan Anorganik = Tim Angsa columns
   - Januari:  cols 2-9
   - Februari: cols 2-9
   - Maret:    cols 2-9
   - April:    cols 2-14 (13 slots) MINUS blue cells at tgl 28-29 (FF00FFFF fill)
   - May:      cols 2-9
   - Juni:     cols 2-9

2. Limbah Sisa Makanan = SOD
   - Maret:  cols 12-13 (SOD RS Pick Up)
   - April:  cols 17-18 (SOD RS) + cols 21-23 (Cator SOD)
   - May:    col 15
   - Juni:   cols 15-17

3. Sampah Lingkungan = Pick Up / Pak Ratian / Daun & Ranting
   - Februari: cols 13-17 (Pick Up UNPAD)
   - Maret:    cols 10-11 (Pick Up)
   - April:    cols 19-20 (Pick Up Seresah)
   - May:      cols 12-14 (Daun & Ranting)
   - Juni:     cols 12-14 (Daun & Ranting) - INCLUDED even though Excel Total omits

4. Sampah Aset = Viar + Mobil Traga (Feb only)
   - Februari: col 12 (Viar) + col 18 (Mobil Traga) - INCLUDED as Aset

5. IPDN separate operator
   - Januari-April: various cols, all zero
   - May: cols 10-11 (only 8 May = 550 kg)
   - Juni: cols 10-11 (zero)

Total per bulan = organik_anorganik + sisa_makanan + lingkungan + aset (UNPAD side) + ipdn.
"""
from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
XLSX = ROOT / "Total Timbulan Sampah 2026  (Bulanan).xlsx"
JSON_PATH = ROOT / "data" / "timbulan.json"
DOCS_JSON_PATH = ROOT / "docs" / "data" / "timbulan.json"

DAY_ID = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
BLUE_FILL = "FF00FFFF"

# Per-sheet layout. 0-indexed column positions in the row tuple.
LAYOUTS = {
    "Januari": {
        "iso": "2026-01", "label": "Januari 2026", "days_in_month": 31,
        "tim_angsa": list(range(2, 10)),
        "sod":       [],
        "lingkungan":[],
        "aset":      [],
        "ipdn":      [10, 11],
        # April is the only sheet with blue-cell handling
        "blue_days": [],
        "blue_cols": [],
    },
    "Februari": {
        "iso": "2026-02", "label": "Februari 2026", "days_in_month": 28,
        "tim_angsa": list(range(2, 10)),
        "sod":       [],
        "lingkungan":[13, 14, 15, 16, 17],
        "aset":      [12, 18],  # Viar + Mobil Traga
        "ipdn":      [10, 11],
        "blue_days": [], "blue_cols": [],
    },
    "Maret": {
        "iso": "2026-03", "label": "Maret 2026", "days_in_month": 31,
        "tim_angsa": list(range(2, 10)),
        "sod":       [12, 13],
        "lingkungan":[10, 11],
        "aset":      [],
        "ipdn":      [],
        "blue_days": [], "blue_cols": [],
    },
    "April": {
        "iso": "2026-04", "label": "April 2026", "days_in_month": 30,
        "tim_angsa": list(range(2, 15)),  # cols 2-14 = Tim Angsa (13 slots)
        "sod":       [17, 18, 21, 22, 23],
        "lingkungan":[19, 20],
        "aset":      [],
        "ipdn":      [15, 16],
        "blue_days": [28, 29],
        "blue_cols": [13, 14],  # cells to exclude on those days
    },
    "May": {
        "iso": "2026-05", "label": "Mei 2026", "days_in_month": 31,
        "tim_angsa": list(range(2, 10)),
        "sod":       [15],
        "lingkungan":[12, 13, 14],
        "aset":      [],
        "ipdn":      [10, 11],
        "blue_days": [], "blue_cols": [],
    },
    "Juni": {
        "iso": "2026-06", "label": "Juni 2026", "days_in_month": 30,
        "tim_angsa": list(range(2, 10)),
        "sod":       [15, 16, 17],
        "lingkungan":[12, 13, 14],  # forced INCLUDE (Excel formula bug)
        "aset":      [],
        "ipdn":      [10, 11],
        "blue_days": [], "blue_cols": [],
    },
    "Juli": {
        "iso": "2026-07", "label": "Juli 2026", "days_in_month": 31,
        "tim_angsa": [], "sod": [], "lingkungan": [], "aset": [], "ipdn": [],
        "blue_days": [], "blue_cols": [],
    },
    "Agustus": {
        "iso": "2026-08", "label": "Agustus 2026", "days_in_month": 31,
        "tim_angsa": [], "sod": [], "lingkungan": [], "aset": [], "ipdn": [],
        "blue_days": [], "blue_cols": [],
    },
    "September": {
        "iso": "2026-09", "label": "September 2026", "days_in_month": 30,
        "tim_angsa": [], "sod": [], "lingkungan": [], "aset": [], "ipdn": [],
        "blue_days": [], "blue_cols": [],
    },
    "Oktober": {
        "iso": "2026-10", "label": "Oktober 2026", "days_in_month": 31,
        "tim_angsa": [], "sod": [], "lingkungan": [], "aset": [], "ipdn": [],
        "blue_days": [], "blue_cols": [],
    },
    "November": {
        "iso": "2026-11", "label": "November 2026", "days_in_month": 30,
        "tim_angsa": [], "sod": [], "lingkungan": [], "aset": [], "ipdn": [],
        "blue_days": [], "blue_cols": [],
    },
    "Desember": {
        "iso": "2026-12", "label": "Desember 2026", "days_in_month": 31,
        "tim_angsa": [], "sod": [], "lingkungan": [], "aset": [], "ipdn": [],
        "blue_days": [], "blue_cols": [],
    },
}


def hari_id(d: dt.date) -> str:
    return DAY_ID[d.weekday()]


def sum_row(row, indexes) -> float:
    tot = 0.0
    for i in indexes:
        if i < len(row):
            v = row[i]
            if isinstance(v, (int, float)):
                tot += v
    return tot


def is_blue(ws_raw, row_idx: int, col_idx: int) -> bool:
    cell = ws_raw.cell(row_idx, col_idx + 1)  # openpyxl is 1-indexed
    fill = cell.fill
    if fill.fill_type != "solid":
        return False
    fg = fill.fgColor.rgb if fill.fgColor else None
    return fg == BLUE_FILL


def extract_month(wb, wb_raw, sheet_name: str, cfg: dict) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    ws_raw = wb_raw[sheet_name]
    iso_prefix = cfg["iso"]
    entries: list[dict] = []
    for row_idx in range(3, ws.max_row + 1):
        row = tuple(ws.cell(row_idx, c + 1).value for c in range(ws.max_column))
        tgl = row[1]
        if not isinstance(tgl, dt.datetime):
            continue
        iso = tgl.strftime("%Y-%m-%d")
        if not iso.startswith(iso_prefix):
            continue

        # Tim Angsa: sum then subtract blue-marked cells for flagged days
        tim_angsa = sum_row(row, cfg["tim_angsa"])
        if tgl.day in cfg["blue_days"]:
            for c in cfg["blue_cols"]:
                if c < len(row) and isinstance(row[c], (int, float)):
                    if is_blue(ws_raw, row_idx, c):
                        tim_angsa -= row[c]

        sod = sum_row(row, cfg["sod"])
        lingkungan = sum_row(row, cfg["lingkungan"])
        aset = sum_row(row, cfg["aset"])
        ipdn = sum_row(row, cfg["ipdn"])

        unpad_total = tim_angsa + sod + lingkungan + aset
        grand_total = unpad_total + ipdn

        if grand_total == 0:
            continue

        entries.append({
            "date": iso,
            "day_of_week": hari_id(tgl.date()),
            "total_kg": round(grand_total, 3),
            "unpad_kg": round(unpad_total, 3),
            "ipdn_kg":  round(ipdn, 3),
            "by_category_kg": {
                "organik_anorganik": round(tim_angsa, 3),
                "sisa_makanan":      round(sod, 3),
                "lingkungan":        round(lingkungan, 3),
                "aset":              round(aset, 3),
            },
            "note": None,
            "quality_flag": None,
        })
    return entries


def build_month_summary(cfg: dict, entries: list[dict]) -> dict:
    iso = cfg["iso"]
    active = [e for e in entries if e["total_kg"] > 0
              and e["quality_flag"] != "excluded_from_average"]
    days_active = len(active)
    dim = cfg["days_in_month"]

    tim_angsa = round(sum(e["by_category_kg"]["organik_anorganik"] for e in entries), 3)
    sod       = round(sum(e["by_category_kg"]["sisa_makanan"]      for e in entries), 3)
    lingkungan= round(sum(e["by_category_kg"]["lingkungan"]        for e in entries), 3)
    aset      = round(sum(e["by_category_kg"]["aset"]              for e in entries), 3)
    unpad_kg  = round(sum(e["unpad_kg"] for e in entries), 3)
    ipdn_kg   = round(sum(e["ipdn_kg"] for e in entries), 3)
    total_kg  = round(unpad_kg + ipdn_kg, 3)
    ipdn_active_days = sum(1 for e in entries if e["ipdn_kg"] > 0)
    has_cat = bool(tim_angsa or sod or lingkungan or aset)

    return {
        "month": iso,
        "label": cfg["label"],
        "total_kg": total_kg if total_kg else None,
        "organik_anorganik_kg": tim_angsa,
        "sisa_makanan_kg":      sod,
        "lingkungan_kg":        lingkungan,
        "aset_kg":              aset,
        "avg_kg_per_calendar_day": round(total_kg / dim, 2) if dim and total_kg else None,
        "days_active": days_active,
        "days_in_month": dim,
        "avg_kg_per_active_day": round(total_kg / days_active, 2) if days_active else None,
        "category_breakdown_available": has_cat,
        "unpad_kg": unpad_kg,
        "ipdn_kg": ipdn_kg,
        "ipdn_active_days": ipdn_active_days,
    }


def main() -> int:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    wb_raw = openpyxl.load_workbook(XLSX)

    # Load existing JSON to preserve top-level fields (unit, vehicle_sources, etc.)
    with JSON_PATH.open("r", encoding="utf-8") as f:
        data = json.load(f)

    # Build daily + monthly from scratch
    daily: list[dict] = []
    monthly: list[dict] = []
    for sheet_name, cfg in LAYOUTS.items():
        entries = extract_month(wb, wb_raw, sheet_name, cfg)
        daily.extend(entries)
        monthly.append(build_month_summary(cfg, entries))

    daily.sort(key=lambda e: e["date"])

    # Update categories metadata (used by dashboard for legends & palette lookup)
    data["categories"] = [
        {"id": "organik_anorganik", "label": "Sampah Organik dan Anorganik",
         "color_key": "organik_anorganik",
         "source": "Truk Tim Angsa"},
        {"id": "sisa_makanan", "label": "Limbah Sisa Makanan",
         "color_key": "sisa_makanan",
         "source": "SOD RS + Cator UNPAD (SOD)"},
        {"id": "lingkungan", "label": "Sampah Lingkungan (Pohon & Ranting)",
         "color_key": "lingkungan",
         "source": "Pick Up / Pak Ratian / Daun & Ranting"},
        {"id": "aset", "label": "Sampah Aset (Meja + Kursi)",
         "color_key": "aset",
         "source": "Viar + Mobil Traga (aset tak terpakai)"},
    ]

    data["daily_entries"] = daily
    data["monthly_summary"] = monthly

    # Update period.end to the last active date
    active_dates = [e["date"] for e in daily if e["total_kg"]]
    if active_dates:
        data["period"]["end"] = max(active_dates)

    data["generated_at"] = dt.date.today().isoformat()

    # Refresh data quality flags for the new scheme
    data["data_quality_flags"] = [
        {"severity": "info",
         "message": "Kategori mengikuti sumber kendaraan: Sampah Organik dan Anorganik (Truk Tim Angsa), Limbah Sisa Makanan (SOD), Sampah Lingkungan (Pick Up / Pak Ratian daun & ranting), dan Sampah Aset (Viar + Mobil Traga untuk meja + kursi bekas)."},
        {"severity": "info",
         "message": "Januari 2026 hanya tercatat 5 hari kerja (26-30 Januari). Rata-rata per hari aktif lebih representatif daripada per hari kalender."},
        {"severity": "warning",
         "message": "April 2026 tanggal 28-29 memiliki cell biru pada kolom 13-14 (13.170 kg) yang tidak dihitung sebagai timbulan sampah (kemungkinan double-entry atau data yang dianulir)."},
        {"severity": "info",
         "message": "Juni 2026 Daun & Ranting (2.630 kg) awalnya terlewat di formula Total Excel; sudah dikoreksi ke kategori Sampah Lingkungan."},
        {"severity": "info",
         "message": "Februari 2026 Viar (467 kg) + Mobil Traga (290 kg) dikategorikan sebagai Sampah Aset — timbulan meja + kursi bekas yang tidak masuk siklus operasional harian."},
    ]

    # Write both copies
    with JSON_PATH.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    with DOCS_JSON_PATH.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # Report
    print(f"[timbulan] wrote {JSON_PATH} + {DOCS_JSON_PATH}")
    print()
    print(f"{'Bulan':10s} | {'Org+Anorg':>10s} | {'SisaMkn':>8s} | {'Lingk':>7s} | {'Aset':>6s} | {'UNPAD':>8s} | {'IPDN':>5s} | {'Total':>8s} | {'HariAkt':>7s}")
    print("-" * 110)
    ytd = {k: 0 for k in ["ta", "sm", "lg", "as", "unpad", "ipdn", "total", "days"]}
    for m in monthly:
        if not m["total_kg"]:
            continue
        print(f"{m['label']:10s} | "
              f"{m['organik_anorganik_kg']:>10.0f} | "
              f"{m['sisa_makanan_kg']:>8.0f} | "
              f"{m['lingkungan_kg']:>7.0f} | "
              f"{m['aset_kg']:>6.0f} | "
              f"{m['unpad_kg']:>8.0f} | "
              f"{m['ipdn_kg']:>5.0f} | "
              f"{m['total_kg']:>8.0f} | "
              f"{m['days_active']:>7d}")
        ytd["ta"] += m["organik_anorganik_kg"]
        ytd["sm"] += m["sisa_makanan_kg"]
        ytd["lg"] += m["lingkungan_kg"]
        ytd["as"] += m["aset_kg"]
        ytd["unpad"] += m["unpad_kg"]
        ytd["ipdn"] += m["ipdn_kg"]
        ytd["total"] += m["total_kg"]
        ytd["days"] += m["days_active"]
    print("-" * 110)
    print(f"{'YTD':10s} | "
          f"{ytd['ta']:>10.0f} | "
          f"{ytd['sm']:>8.0f} | "
          f"{ytd['lg']:>7.0f} | "
          f"{ytd['as']:>6.0f} | "
          f"{ytd['unpad']:>8.0f} | "
          f"{ytd['ipdn']:>5.0f} | "
          f"{ytd['total']:>8.0f} | "
          f"{ytd['days']:>7d}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
