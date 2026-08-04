"""Parse Data Listrik.xlsx -> electricity.json (Jatinangor saja).

Sumber: satu sheet, blok lebar. Baris 4 = header, baris 5 = penanda tahun,
baris 6-17 = Januari..Desember. Kolom Jatinangor per tahun:
  J(10)=2022, N(14)=2023, R(18)=2024, V(22)=2025, Z(26)=2026.

Blok kiri (kolom B) hanya 2024 dan identik dengan kolom R — dipakai sebagai
cek-silang, bukan sumber. Blok tinggi (baris 42+) memuat seri berbeda
(nilai 2022 lebih tinggi, kemungkinan termasuk RSPTN) — TIDAK dipakai.

Keluaran: seri bulanan kontinu 2022-01 .. bulan terakhir, tiap entri membawa
kWh dan rata-rata bergerak 12-bulan, plus ringkasan per tahun.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _utils import (
    base_dataset,
    find_project_root,
    find_source,
    flag,
    resolve_output,
    round_to,
    write_json,
)

SOURCE_XLSX = "Data Listrik.xlsx"
DATASET_ID = "electricity"
LOKASI = "Jatinangor"
SATUAN = "kWh"

# Kolom Jatinangor per tahun di blok lebar (baris header = 4, penanda tahun = 5).
KOLOM_JATINANGOR = {2022: 10, 2023: 14, 2024: 18, 2025: 22, 2026: 26}
COL_BULAN = 9          # kolom I
BARIS_HEADER = 4
BARIS_TAHUN = 5
BARIS_DATA = range(6, 18)   # Jan..Des

BULAN_ID = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]
BULAN_SINGKAT = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun",
                 "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
MA_WINDOW = 12          # rata-rata bergerak 12 bulan (menyerap musiman tahunan)


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, AttributeError):
        return None


def bulan_ke_indeks(teks: str) -> int | None:
    """'Januari' -> 0. Kembalikan None kalau bukan nama bulan yang dikenal."""
    t = str(teks or "").strip().lower()
    for i, nama in enumerate(BULAN_ID):
        if t == nama.lower() or t == nama[:3].lower():
            return i
    return None


def main() -> int:
    ap = argparse.ArgumentParser(description=f"Build {DATASET_ID}.json (Jatinangor)")
    ap.add_argument("--source", default=None)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("[parse_electricity] openpyxl belum terpasang: pip install openpyxl", file=sys.stderr)
        return 1

    project = find_project_root()
    src = Path(args.source) if args.source else (find_source(SOURCE_XLSX) or project / SOURCE_XLSX)
    if not src.exists():
        print(f"[parse_electricity] sumber tidak ditemukan: {src}", file=sys.stderr)
        print(f"[parse_electricity] dicari juga di '<Dashboard>/Data dan Pengetahuan/'.", file=sys.stderr)
        return 1
    print(f"[parse_electricity] sumber: {src}")

    ws = openpyxl.load_workbook(src, data_only=True)["Sheet1"]

    # --- Verifikasi tata letak lewat LABEL, bukan asumsi buta. Kalau berubah, gagal-keras.
    label_tahun_asli: dict[int, str] = {}
    for tahun, kol in KOLOM_JATINANGOR.items():
        hdr = str(ws.cell(BARIS_HEADER, kol).value or "").strip()
        penanda = str(ws.cell(BARIS_TAHUN, kol).value or "").strip()
        if "jatinangor" not in hdr.lower():
            print(
                f"[parse_electricity] GAGAL: kolom {kol} untuk {tahun} berlabel '{hdr}', "
                f"bukan 'Jatinangor'. Tata letak Excel berubah; perbarui KOLOM_JATINANGOR.",
                file=sys.stderr,
            )
            return 1
        if str(tahun) not in penanda:
            print(
                f"[parse_electricity] GAGAL: penanda tahun kolom {kol} = '{penanda}', "
                f"tidak memuat '{tahun}'. Tata letak berubah.",
                file=sys.stderr,
            )
            return 1
        label_tahun_asli[tahun] = hdr

    # --- Baca nilai per bulan per tahun
    tahun_urut = sorted(KOLOM_JATINANGOR)
    nilai: dict[tuple[int, int], float] = {}   # (tahun, indeks_bulan) -> kWh
    for r in BARIS_DATA:
        idx = bulan_ke_indeks(ws.cell(r, COL_BULAN).value)
        if idx is None:
            continue
        for tahun, kol in KOLOM_JATINANGOR.items():
            v = num(ws.cell(r, kol).value)
            # 0 dan kosong sama-sama berarti "belum ada data" untuk listrik bulanan.
            if v is not None and v > 0:
                nilai[(tahun, idx)] = round_to(v, 2)

    if not nilai:
        print("[parse_electricity] tidak ada nilai terbaca", file=sys.stderr)
        return 1

    # --- Seri bulanan kontinu dari 2022-01 sampai bulan terisi terakhir
    tahun_min, tahun_maks = tahun_urut[0], tahun_urut[-1]
    bulan_terakhir = max(idx for (th, idx) in nilai if th == tahun_maks)

    monthly: list[dict] = []
    for tahun in range(tahun_min, tahun_maks + 1):
        batas = bulan_terakhir if tahun == tahun_maks else 11
        for idx in range(0, batas + 1):
            monthly.append({
                "month": f"{tahun}-{idx + 1:02d}",
                "label": f"{BULAN_SINGKAT[idx]} {tahun}",
                "year": tahun,
                "month_name": BULAN_ID[idx],
                "kwh": nilai.get((tahun, idx)),   # None kalau bolong di tengah
            })

    # --- Rata-rata bergerak 12 bulan (trailing). Butuh 12 nilai non-null berurutan.
    seri = [m["kwh"] for m in monthly]
    for i, m in enumerate(monthly):
        jendela = seri[max(0, i - MA_WINDOW + 1): i + 1]
        valid = [x for x in jendela if x is not None]
        m["ma12"] = round_to(sum(valid) / len(valid), 2) if len(valid) == MA_WINDOW else None

    # --- Ringkasan per tahun
    yearly: list[dict] = []
    for tahun in tahun_urut:
        vals = [nilai[(tahun, i)] for i in range(12) if (tahun, i) in nilai]
        if not vals:
            continue
        yearly.append({
            "year": tahun,
            "months": len(vals),
            "total_kwh": round_to(sum(vals), 2),
            "avg_kwh": round_to(sum(vals) / len(vals), 2),
            "label_sumber": label_tahun_asli[tahun],
            "lengkap": len(vals) == 12,
        })

    total_kwh = round_to(sum(m["kwh"] for m in monthly if m["kwh"] is not None), 2)
    terisi = [m for m in monthly if m["kwh"] is not None]

    # --- Flags: definisi berubah, catatan RSPTN, tahun parsial, bolong di tengah
    flags = [
        flag("info", f"Hanya data {LOKASI}; kolom Pangandaran, Arjasari, dan DU diabaikan sesuai permintaan."),
        flag("info", f"Catatan sumber: '{LOKASI} tanpa RSPTN' (konsumsi rumah sakit tidak termasuk)."),
    ]
    ganti_def = sorted(th for th, lab in label_tahun_asli.items() if "just unpad" in lab.lower())
    if ganti_def:
        flags.append(flag(
            "warning",
            f"Definisi berubah: tahun {', '.join(map(str, ganti_def))} berlabel "
            f"'Jatinangor (Just UNPAD)' di sumber, sedangkan tahun sebelumnya 'Jatinangor'. "
            f"Angka lintas tahun mungkin tidak sepenuhnya setara.",
        ))
    for y in yearly:
        if not y["lengkap"]:
            flags.append(flag(
                "warning",
                f"Tahun {y['year']} baru {y['months']} bulan (parsial); total dan rata-ratanya belum setahun penuh.",
            ))
    bolong = [m["month"] for m in monthly if m["kwh"] is None]
    if bolong:
        flags.append(flag(
            "warning",
            f"{len(bolong)} bulan kosong di tengah rentang: {', '.join(bolong)}.",
        ))

    data = base_dataset(
        DATASET_ID,
        source_files=[SOURCE_XLSX],
        period={"start": monthly[0]["month"] + "-01", "end": terisi[-1]["month"] + "-01"},
    )
    data["data_quality_flags"] = flags
    data["location"] = LOKASI
    data["unit"] = SATUAN
    data["ma_window_months"] = MA_WINDOW
    data["years"] = tahun_urut
    data["summary"] = {
        "total_kwh": total_kwh,
        "months_with_data": len(terisi),
        "year_count": len(yearly),
        "first_month": monthly[0]["month"],
        "last_month": terisi[-1]["month"],
        "avg_kwh_per_month": round_to(total_kwh / len(terisi), 2) if terisi else 0,
    }
    data["yearly"] = yearly
    data["monthly"] = monthly

    out_dir = resolve_output(args.out)
    target = out_dir / f"{DATASET_ID}.json"
    write_json(target, data)
    print(
        f"[parse_electricity] wrote {target} "
        f"({len(terisi)} bulan berisi, {len(yearly)} tahun, "
        f"{total_kwh:,.0f} kWh total, {len(flags)} flags)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
