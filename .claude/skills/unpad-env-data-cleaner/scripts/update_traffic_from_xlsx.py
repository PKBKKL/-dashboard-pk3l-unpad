"""Rebuild data/traffic_accidents.json from Kecelakaan Lalu Lintas.xlsx.

Reads two sheets ("2025" and "2026") and:
- Main table: rows 4-11, monthly counts per jenis kecelakaan
- Location table (sheet 2026 only): rows 4+, cols 15-18 (No, Jenis, Lokasi, Jumlah)

Preserves period.start and other top-level fields; recomputes period.end from
the last active month.
"""
from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
XLSX = ROOT / "Kecelakaan Lalu Lintas.xlsx"
JSON_PATH = ROOT / "data" / "traffic_accidents.json"
DOCS_JSON_PATH = ROOT / "docs" / "data" / "traffic_accidents.json"

# Row 4 -> Tunggal Sepeda Motor, Row 5 -> Tunggal Mobil, ...
# We match by the label at col 1 rather than fixed row index because
# 2026 introduced a new row ("Tabrakan antar roda empat dan beam").
TYPE_MAP = {
    "Tunggal Sepeda Motor":                    "tunggal_motor",
    "Tunggal Mobil":                           "tunggal_mobil",
    "Beam":                                    "beam",
    "Tabrakan antar roda dua":                 "tabrak_2roda",
    "Tabrakan antar roda dua dan beam":        "tabrak_2roda_beam",
    "Tabrakan antar roda dua dan roda empat":  "tabrak_2roda_4roda",
    "Tabrakan antar roda empat dan beam":      "tabrak_4roda_beam",
    "Pejalan kaki":                            "pejalan_kaki",
}

# Location label at col 15 sub-header row (year-specific)
LOCATION_ID_MAP = {
    "Pertanian":        "faperta",
    "Tugu makalangan":  "tugu-makalangan",
    "FKG":              "fkg",
    "FEB":              "feb",
}

MONTHS_ID = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
             "Juli", "Agustus", "September", "Oktober", "November", "Desember"]


def parse_year_sheet(ws, year: int) -> dict:
    """Return {'monthly': [...], 'total_yearly_computed': N}.

    Layout:
      row 3: month labels at col B onwards
      rows 4..N: jenis label at col A, monthly counts at col B onwards
      last data row: "Total Kasus" label (skipped, we recompute)
    """
    # Discover column -> ISO month map from row 3
    col_to_month = {}
    for col in range(2, ws.max_column + 1):
        v = ws.cell(3, col).value
        if isinstance(v, str) and v in MONTHS_ID:
            month_idx = MONTHS_ID.index(v) + 1
            col_to_month[col] = f"{year}-{month_idx:02d}"

    # Aggregate rows into by-month totals
    by_month: dict[str, dict[str, int]] = {}
    for row in range(4, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if not isinstance(label, str):
            continue
        if label.strip().lower().startswith("total"):
            break
        type_id = TYPE_MAP.get(label.strip())
        if not type_id:
            continue
        for col, iso in col_to_month.items():
            v = ws.cell(row, col).value
            if isinstance(v, (int, float)) and v > 0:
                by_month.setdefault(iso, {})[type_id] = \
                    by_month.setdefault(iso, {}).get(type_id, 0) + int(v)

    monthly = []
    for iso in sorted(by_month.keys()):
        types = by_month[iso]
        monthly.append({
            "month": iso,
            "by_type": types,
            "total": sum(types.values()),
        })

    total = sum(m["total"] for m in monthly)
    return {"monthly": monthly, "total_yearly_computed": total}


def parse_location_table_2026(ws) -> list[dict]:
    """Extract 'Berdasarkan Lokasi' details (sheet 2026, cols 15-18).

    Structure:
      row 3: header ["No", "Jenis Kecelakaan", "Lokasi", "Jumlah"]
      row 4: month section header (e.g. "Januari" at col 15)
      row 5..N: rows with No | Jenis | Lokasi | Jumlah
    """
    incidents = []
    current_month_iso: str | None = None
    no_counter = 0
    for row in range(4, ws.max_row + 1):
        col_no = ws.cell(row, 16).value       # 1-indexed col 16 == 0-indexed 15
        col_jenis = ws.cell(row, 17).value
        col_lokasi = ws.cell(row, 18).value
        col_jml = ws.cell(row, 19).value

        # Section header row: col15 is a month name, others empty
        if isinstance(col_no, str) and col_no in MONTHS_ID and col_jenis is None:
            month_idx = MONTHS_ID.index(col_no) + 1
            current_month_iso = f"2026-{month_idx:02d}"
            continue

        # Data row: col15 is a number, cols 16/17/18 filled
        if isinstance(col_no, (int, float)) and col_jenis and col_lokasi:
            no_counter += 1
            jenis_raw = str(col_jenis).strip()
            lokasi_raw = str(col_lokasi).strip()
            # Match jenis to type_id (fuzzy — location table uses shorter labels)
            type_id = _match_type_id(jenis_raw)
            location_id = LOCATION_ID_MAP.get(lokasi_raw, lokasi_raw.lower().replace(" ", "-"))
            entry = {
                "no": no_counter,
                "month": current_month_iso or "2026-01",
                "type": type_id,
                "location_id": location_id,
                "location_label_raw": lokasi_raw,
                "count": int(col_jml) if isinstance(col_jml, (int, float)) else 1,
            }
            # Preserve original note if jenis_raw doesn't match a standard label
            if type_id is None or jenis_raw.lower() not in [t.lower() for t in TYPE_MAP]:
                entry["note"] = jenis_raw
            incidents.append(entry)
    return incidents


def _match_type_id(jenis_raw: str) -> str | None:
    """Fuzzy-match location table's shorter labels to main type IDs."""
    j = jenis_raw.lower()
    if "beam" in j and ("mobil" in j or "empat" in j):
        return "tabrak_4roda_beam"
    if "beam" in j and ("dua" in j or "motor" in j):
        return "tabrak_2roda_beam"
    if "antar roda dua" in j or ("roda dua" in j and "empat" not in j and "beam" not in j):
        return "tabrak_2roda"
    if "roda dua" in j and "empat" in j:
        return "tabrak_2roda_4roda"
    if "tunggal" in j and "mobil" in j:
        return "tunggal_mobil"
    if "tunggal" in j and ("motor" in j or "sepeda motor" in j):
        return "tunggal_motor"
    if "pejalan" in j:
        return "pejalan_kaki"
    if j.strip() == "beam":
        return "beam"
    return None


def main() -> int:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    with JSON_PATH.open("r", encoding="utf-8") as f:
        data = json.load(f)

    # Rebuild yearly for both 2025 and 2026
    yearly = []
    for year in [2025, 2026]:
        sheet_name = str(year)
        if sheet_name not in wb.sheetnames:
            continue
        parsed = parse_year_sheet(wb[sheet_name], year)
        entry = {
            "year": year,
            "monthly": parsed["monthly"],
            "total_yearly_computed": parsed["total_yearly_computed"],
            "total_yearly_reported": parsed["total_yearly_computed"],
        }
        if year == 2026 and parsed["monthly"]:
            entry["ytd_through_month"] = parsed["monthly"][-1]["month"]
        yearly.append(entry)
    data["yearly"] = yearly

    # Rebuild incidents_detail_2026 from location table
    if "2026" in wb.sheetnames:
        data["incidents_detail_2026"] = parse_location_table_2026(wb["2026"])

    # Update period.end to the last active month in 2026 (or 2025 fallback)
    last_iso = None
    for e in yearly:
        for m in e["monthly"]:
            if m["total"] > 0 and (last_iso is None or m["month"] > last_iso):
                last_iso = m["month"]
    if last_iso:
        y, mo = map(int, last_iso.split("-"))
        # end of month
        next_month = dt.date(y + (1 if mo == 12 else 0),
                             1 if mo == 12 else mo + 1, 1)
        data["period"]["end"] = (next_month - dt.timedelta(days=1)).isoformat()

    data["generated_at"] = dt.date.today().isoformat()

    with JSON_PATH.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    with DOCS_JSON_PATH.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"[traffic] wrote {JSON_PATH} + {DOCS_JSON_PATH}")
    print()
    for e in yearly:
        print(f"Year {e['year']}: total = {e['total_yearly_computed']}")
        for m in e["monthly"]:
            print(f"  {m['month']}: {m['total']:>2}  {m['by_type']}")
    if data.get("incidents_detail_2026"):
        print()
        print(f"Location detail 2026: {len(data['incidents_detail_2026'])} entries")

    return 0


if __name__ == "__main__":
    sys.exit(main())
