"""Bangun traffic_accidents.json dari workbook kecelakaan.

Menggantikan parse_traffic_accidents.py (yang membaca MD lama; MD itu hanya
memuat 4 kasus untuk 2026, sementara workbook induk memuat 10). Self-contained:
tidak membaca JSON lama, dan menulis lewat --out sehingga tunduk pada staging +
anti-regresi di run_all.py.

TAHUN TIDAK DI-HARDCODE. Setiap sheet yang namanya empat digit angka
diperlakukan sebagai satu tahun. Menambah 2027 cukup dengan menambah sheet
bernama '2027' di workbook — tanpa menyentuh kode ini.

Sheet tahun:
  baris 1  : 'Jenis Kecelakaan' (kol A), 'Tahun YYYY' (kol B)
  baris 3  : nama bulan mulai kolom B
  baris 4+ : label jenis di kolom A, jumlah per bulan di kolom B+
  baris 'Total Kasus' : pembacaan berhenti di sini, total dihitung ulang
  kolom 16-19 (opsional) : tabel 'Berdasarkan Lokasi' (No, Jenis, Lokasi, Jumlah)

Output memakai field `incidents_detail` (bukan `incidents_detail_2026`), dengan
kolom `year` di tiap barisnya — spec v1.1.

Butuh openpyxl.
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _utils import base_dataset, find_source, flag, resolve_output, write_json

try:
    import openpyxl
except ImportError:  # pragma: no cover
    print("[parse_traffic_accidents_xlsx] butuh openpyxl: python -m pip install openpyxl", file=sys.stderr)
    raise

# Workbook induk lebih dulu. Workbook lama di 'Data dan Pengetahuan' (6 Apr 2026)
# hanya memuat 4 kasus 2026; Mei (3) & Juni (3) tidak ada di sana.
SOURCE_XLSX = "Kecelakaan Lalu Lintas (MASTER).xlsx"
FALLBACK_XLSX = "Kecelakaan Lalu Lintas.xlsx"
DATASET_ID = "traffic_accidents"
YEAR_SHEET = re.compile(r"^\d{4}$")

MONTHS_ID = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
             "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

TYPE_MAP = {
    "Tunggal Sepeda Motor": "tunggal_motor",
    "Tunggal Mobil": "tunggal_mobil",
    "Beam": "beam",
    "Tabrakan antar roda dua": "tabrak_2roda",
    "Tabrakan antar roda dua dan beam": "tabrak_2roda_beam",
    "Tabrakan antar roda dua dan roda empat": "tabrak_2roda_4roda",
    "Tabrakan antar roda empat dan beam": "tabrak_4roda_beam",
    "Pejalan kaki": "pejalan_kaki",
}

VEHICLE_TYPES = [
    {"id": "tunggal_motor", "label": "Tunggal Sepeda Motor"},
    {"id": "tunggal_mobil", "label": "Tunggal Mobil"},
    {"id": "beam", "label": "Beam (sepeda listrik)"},
    {"id": "tabrak_2roda", "label": "Tabrakan antar roda dua"},
    {"id": "tabrak_2roda_beam", "label": "Tabrakan roda dua dan Beam"},
    {"id": "tabrak_2roda_4roda", "label": "Tabrakan roda dua dan roda empat"},
    {"id": "tabrak_4roda_beam", "label": "Tabrakan roda empat dan Beam"},
    {"id": "pejalan_kaki", "label": "Pejalan kaki"},
]

LOCATION_ID_MAP = {
    "Pertanian": "faperta",
    "Tugu makalangan": "tugu-makalangan",
    "FKG": "fkg",
    "FEB": "feb",
}

PERIOD_START = "2025-04-01"


def _match_type_id(jenis_raw: str) -> str | None:
    j = jenis_raw.lower()
    if "beam" in j and ("mobil" in j or "empat" in j):
        return "tabrak_4roda_beam"
    if "beam" in j and ("dua" in j or "motor" in j):
        return "tabrak_2roda_beam"
    if "roda dua" in j and "empat" in j:
        return "tabrak_2roda_4roda"
    if "antar roda dua" in j or ("roda dua" in j and "empat" not in j and "beam" not in j):
        return "tabrak_2roda"
    if "tunggal" in j and "mobil" in j:
        return "tunggal_mobil"
    if "tunggal" in j and "motor" in j:
        return "tunggal_motor"
    if "pejalan" in j:
        return "pejalan_kaki"
    if j.strip() == "beam":
        return "beam"
    return None


def parse_year_sheet(ws, year: int) -> dict:
    col_to_month = {}
    for col in range(2, ws.max_column + 1):
        v = ws.cell(3, col).value
        if isinstance(v, str) and v.strip() in MONTHS_ID:
            col_to_month[col] = f"{year}-{MONTHS_ID.index(v.strip()) + 1:02d}"

    by_month: dict[str, dict[str, int]] = {}
    for row in range(4, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if not isinstance(label, str):
            continue
        if label.strip().lower().startswith("total"):
            break
        type_id = TYPE_MAP.get(label.strip())
        if not type_id:
            continue
        for col, iso in col_to_month.items():
            v = ws.cell(row, col).value
            if isinstance(v, (int, float)) and v > 0:
                by_month.setdefault(iso, {})
                by_month[iso][type_id] = by_month[iso].get(type_id, 0) + int(v)

    monthly = [{"month": iso, "by_type": by_month[iso], "total": sum(by_month[iso].values())}
               for iso in sorted(by_month)]
    return {"monthly": monthly, "total_yearly_computed": sum(m["total"] for m in monthly)}


def parse_location_table(ws, year: int) -> list[dict]:
    """Tabel 'Berdasarkan Lokasi' di kolom 16-19. Opsional; sheet tanpa tabel ini
    mengembalikan daftar kosong. Berlaku untuk tahun apa pun, bukan hanya 2026."""
    incidents: list[dict] = []
    current_month = None
    no_counter = 0
    for row in range(4, ws.max_row + 1):
        col_no = ws.cell(row, 16).value
        col_jenis = ws.cell(row, 17).value
        col_lokasi = ws.cell(row, 18).value
        col_jml = ws.cell(row, 19).value

        if isinstance(col_no, str) and col_no.strip() in MONTHS_ID and col_jenis is None:
            current_month = f"{year}-{MONTHS_ID.index(col_no.strip()) + 1:02d}"
            continue

        if isinstance(col_no, (int, float)) and col_jenis and col_lokasi:
            no_counter += 1
            jenis_raw = str(col_jenis).strip()
            lokasi_raw = str(col_lokasi).strip()
            type_id = _match_type_id(jenis_raw)
            entry = {
                "no": no_counter,
                "year": year,
                "month": current_month or f"{year}-01",
                "type": type_id,
                "location_id": LOCATION_ID_MAP.get(lokasi_raw, lokasi_raw.lower().replace(" ", "-")),
                "location_label_raw": lokasi_raw,
                "count": int(col_jml) if isinstance(col_jml, (int, float)) else 1,
            }
            if type_id is None or jenis_raw.lower() not in [t.lower() for t in TYPE_MAP]:
                entry["note"] = jenis_raw
            incidents.append(entry)
    return incidents


def _active_months(yearly: list[dict]) -> list[str]:
    return sorted(m["month"] for e in yearly for m in e["monthly"] if m["total"] > 0)


def _period_start(yearly: list[dict]) -> str | None:
    months = _active_months(yearly)
    if not months:
        return None
    y, mo = map(int, months[0].split("-"))
    return dt.date(y, mo, 1).isoformat()


def _period_end(yearly: list[dict]) -> str | None:
    months = _active_months(yearly)
    if not months:
        return None
    y, mo = map(int, months[-1].split("-"))
    nxt = dt.date(y + (1 if mo == 12 else 0), 1 if mo == 12 else mo + 1, 1)
    return (nxt - dt.timedelta(days=1)).isoformat()


def _warn_if_fallback_is_newer(master: Path) -> None:
    """Kalau pemilik mengetik data baru di workbook lama, editannya akan
    DIABAIKAN diam-diam karena parser membaca workbook induk. Berisik lebih baik."""
    old = find_source(FALLBACK_XLSX)
    if not old or not old.exists() or old.name == master.name:
        return
    if old.stat().st_mtime > master.stat().st_mtime:
        print(
            f"\n[parse_{DATASET_ID}] PERINGATAN KERAS\n"
            f"  '{old.name}' lebih baru daripada '{master.name}'.\n"
            f"  Parser membaca workbook induk, jadi editan Anda di '{old.name}' TIDAK terbaca.\n"
            f"  Pindahkan perubahannya ke workbook induk, atau sebutkan --source secara eksplisit.\n",
            file=sys.stderr,
        )


def main() -> int:
    ap = argparse.ArgumentParser(description=f"[PENSIUN sebagai parser] Build {DATASET_ID}.json dari XLSX")
    ap.add_argument("--source", default=None, help=f"Path ke {SOURCE_XLSX}")
    ap.add_argument("--out", default=None)
    ap.add_argument("--i-know-this-is-retired", action="store_true")
    args = ap.parse_args()

    if not args.i_know_this_is_retired:
        print(
            "[parse_traffic_accidents_xlsx] SKRIP INI PENSIUN sebagai parser dan menolak jalan.\n"
            "  Fungsi pembacaannya masih dipakai import_inbox.py sebagai pustaka.\n"
            "  Sumber resmi sekarang: data/_ledger/traffic_accidents.csv\n"
            "  Pakai:  python parse_traffic_accidents_ledger.py --out data\n"
            "  Tambah data baru:  python import_inbox.py --dataset kecelakaan",
            file=sys.stderr,
        )
        return 1

    if args.source:
        src = Path(args.source)
    else:
        src = find_source(SOURCE_XLSX) or find_source(FALLBACK_XLSX)
    if not src or not src.exists():
        print(f"[parse_{DATASET_ID}] sumber '{SOURCE_XLSX}' (atau '{FALLBACK_XLSX}') tidak ditemukan "
              f"di root repo maupun '<Dashboard>/Data dan Pengetahuan/'.\n"
              f"ABORT — {DATASET_ID}.json TIDAK diubah.", file=sys.stderr)
        return 1
    if src.name == FALLBACK_XLSX:
        print(f"[parse_{DATASET_ID}] PERINGATAN: memakai workbook lama '{FALLBACK_XLSX}'. "
              f"Ia tidak memuat Mei & Juni 2026. Pengaman anti-regresi akan memblokir promosi.",
              file=sys.stderr)

    _warn_if_fallback_is_newer(src)

    wb = openpyxl.load_workbook(src, data_only=True)

    years = sorted(int(s) for s in wb.sheetnames if YEAR_SHEET.match(s))
    if not years:
        print(f"[parse_{DATASET_ID}] {src.name} tidak punya satu pun sheet bernama tahun "
              f"(mis. '2025'). Sheet yang ada: {wb.sheetnames}. ABORT.", file=sys.stderr)
        return 1
    latest = years[-1]

    yearly: list[dict] = []
    detail: list[dict] = []
    for year in years:
        ws = wb[str(year)]
        parsed = parse_year_sheet(ws, year)
        entry = {
            "year": year,
            "monthly": parsed["monthly"],
            "total_yearly_computed": parsed["total_yearly_computed"],
            "total_yearly_reported": parsed["total_yearly_computed"],
        }
        if year == latest and parsed["monthly"]:
            entry["ytd_through_month"] = parsed["monthly"][-1]["month"]
        yearly.append(entry)
        detail.extend(parse_location_table(ws, year))

    if all(e["total_yearly_computed"] == 0 for e in yearly):
        print(f"[parse_{DATASET_ID}] tidak ada satu pun kasus terbaca dari {src.name}. "
              f"Menolak menulis output kosong. ABORT.", file=sys.stderr)
        return 1

    wb.close()

    start = _period_start(yearly) or PERIOD_START
    data = base_dataset(
        DATASET_ID,
        source_files=[src.name],
        period={"start": start, "end": _period_end(yearly) or start},
    )
    data["generated_at"] = dt.date.today().isoformat()
    data["data_quality_flags"] = [
        flag("info", "Persentase di sumber dihitung terhadap total tahun (33 untuk 2025). "
                     "Untuk 2026 yang masih parsial, persentase dihitung ulang di frontend "
                     "berdasarkan total YTD."),
    ]
    data["vehicle_types"] = VEHICLE_TYPES
    data["yearly"] = yearly
    data["incidents_detail"] = detail

    out_dir = resolve_output(args.out)
    target = out_dir / f"{DATASET_ID}.json"
    write_json(target, data)
    print(f"[parse_{DATASET_ID}] sumber: {src}")
    print(f"[parse_{DATASET_ID}] tahun terbaca dari sheet: {years}")
    print(f"[parse_{DATASET_ID}] wrote {target} (" +
          ", ".join(f"{e['year']}: {e['total_yearly_computed']}" for e in yearly) +
          f", {len(detail)} detail lokasi)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
